from openpyxl.styles.borders import Border, Side, BORDER_THIN
import json
import os
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser
from openpyxl import Workbook
from openpyxl.chart import Reference, PieChart3D
from openpyxl.styles import Font
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import io
import urllib3
from openpyxl import load_workbook


def setOverviewPage(workSheetOverView, excelData):
    data = json.loads(excelData['Overview'])
    PlanDetail = json.loads(excelData['PlanDetail'])
    verticalCount = excelData['verticalCount']
    OverviewExtraPoints = json.loads(excelData['OverviewExtraPoints'])
    CountryCovered = json.loads(excelData['CountryCovered'])

    # workSheetOverView.merge_cells('A7:E7')
    # UserNameCell = workSheetOverView.cell(row=7, column=1)
    # UserNameCell.alignment = Alignment(
    #     horizontal="center", vertical="center", wrapText=True)
    # UserNameCell.font = Font(bold=True, size=16)
    # UserNameCell.value = 'Global Media Kit'

    workSheetOverView['C6']

    for row in PlanDetail:
        workSheetOverView.append(row)

    workSheetOverView.merge_cells('A14:C14')

    workSheetOverView['C14']
    for row in data:
        workSheetOverView.append(row)
    pie = PieChart3D()
    labels = Reference(workSheetOverView, min_col=1,
                       min_row=16, max_row=7+16)
    data = Reference(workSheetOverView, min_col=2,
                     min_row=15, max_row=7+15)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Budget allocation"
    pie.height = 5.6  # default is 7.5
    pie.width = 11  # default is 15
    workSheetOverView.add_chart(pie, "E14")

    http = urllib3.PoolManager()
    r = http.request(
        'GET', 'https://globalmediakit.com/assets/images/logo.png')
    image_file = io.BytesIO(r.data)
    img = Image(image_file)
    img.height = 75
    img.width = 250
    workSheetOverView.add_image(img, 'A2')

    setSheet(workSheetOverView)
    workSheetOverView['A15'].alignment = Alignment(wrap_text=True)
    workSheetOverView['A15'].font = Font(bold=True, size=11)
    workSheetOverView['A15'].fill = PatternFill("solid", fgColor="f7941d")
    workSheetOverView['B15'].alignment = Alignment(wrap_text=True)
    workSheetOverView['B15'].font = Font(bold=True, size=11)
    workSheetOverView['B15'].fill = PatternFill("solid", fgColor="f7941d")
    workSheetOverView['C15'].alignment = Alignment(wrap_text=True)
    workSheetOverView['C15'].font = Font(bold=True, size=11)
    workSheetOverView['C15'].fill = PatternFill("solid", fgColor="f7941d")
    workSheetOverView['A23'].font = Font(bold=True, size=11)
    workSheetOverView['B23'].font = Font(bold=True, size=11)
    workSheetOverView['C23'].font = Font(bold=True, size=11)

    
    i = 27
    for row in OverviewExtraPoints:
        for cell in workSheetOverView[str(i)]:
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrapText=True)
        for item in row:

            workSheetOverView.merge_cells('A'+str(i)+':H'+str(i)+'')
            workSheetOverView.cell(row=i, column=1).value = item
            workSheetOverView.cell(
                row=i, column=1).alignment = Alignment(wrapText=True)
            i = i+1
    workSheetOverView.cell(
        row=i-1, column=1).hyperlink = "https://www.globalmediakit.com"
        
    i = 12
    column = 1
    # workSheetOverView.merge_cells('J' + str(i) + ':J' + str(i))
    workSheetOverView.cell(row=i, column=column).value = "Countries Covered"
    workSheetOverView.cell(
        row=i, column=column).font = Font(bold=True, size=11)
    # workSheetOverView.cell(row=i, column=column).border = Border(
    #     left=Side(border_style=BORDER_THIN, color='000000'),
    #     right=Side(border_style=BORDER_THIN, color='000000'),
    #     top=Side(border_style=BORDER_THIN, color='000000'),
    #     bottom=Side(border_style=BORDER_THIN, color='000000')
    # )
    # i = i+1
    # for Country in CountryCovered:
    workSheetOverView.merge_cells('B' + str(i) + ':I' + str(i))
    workSheetOverView.cell(row=i, column=2).value = CountryCovered
    # workSheetOverView.cell(row=i, column=2).border = Border(
    #     left=Side(border_style=BORDER_THIN, color='000000'),
    #     right=Side(border_style=BORDER_THIN, color='000000'),
    #     top=Side(border_style=BORDER_THIN, color='000000'),
    #     bottom=Side(border_style=BORDER_THIN, color='000000')
    # )
    #     i = i+1
    setAutoWidth(workSheetOverView)

    # workSheetOverView.merge_cells('J14:M26')
    # workSheetOverView.cell(row=14, column=10).value = "	Delhi, Mumbai, Gujarat, Pune, Chennai, Ahmedabad, Kolkata, Chandigarh, Lucknow, Kochi, Bhubaneswa"

    headingCell = workSheetOverView.cell(row=14, column=1)
    headingCell.alignment = Alignment(
        horizontal="center", vertical="center", wrapText=True)
    headingCell.font = Font(bold=True, size=14)
    headingCell.value = 'Plan summary'
    headingCell.border = Border(
        left=Side(border_style=None, color='FFFFFFFF'),
        right=Side(border_style=None, color='FFFFFFFF'),
        top=Side(border_style=None, color='FFFFFFFF'),
        bottom=Side(border_style=None, color='FFFFFFFF'),
    )
    headingCell.alignment = Alignment(
        horizontal="center", vertical="center", wrapText=True)
    return workSheetOverView


def setVerticalPage(workBook, data):

    for row in data:
        workSheetVertical = row['VerticalName']
        workBook.create_sheet(workSheetVertical)

        workBook[workSheetVertical].freeze_panes = workBook[workSheetVertical]['E1']
        # workBook[workSheetVertical].merge_cells('A3:C3')

        workBook[workSheetVertical]['C5']
        workBook[workSheetVertical].cell(2, 3)
        for sheetRow in row['Data']:
            workBook[workSheetVertical].append(sheetRow)
        i = 7
        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='000000'),
            right=Side(border_style=BORDER_THIN, color='000000'),
            top=Side(border_style=BORDER_THIN, color='000000'),
            bottom=Side(border_style=BORDER_THIN, color='000000')
        )
        j=0
        for sheetRow in row['Data']:
            for cell in workBook[workSheetVertical][str(i)]:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrapText=True)
                if cell.value or cell.value == 0:
                    cell.border = thin_border
            if(row['MediaURL'][j]):
                workBook[workSheetVertical]['A'+str(i-1)].hyperlink = row['MediaURL'][j]
            j=j+1
            if(workBook[workSheetVertical]['E'+str(i)].value and workBook[workSheetVertical]['E'+str(i)].value!='N/A'):
                value = workBook[workSheetVertical]['E'+str(i)].value
                workBook[workSheetVertical]['E'+str(i)].value = 'View Media Kit'
                workBook[workSheetVertical]['E'+str(i)].hyperlink = value
                # if(row['VerticalName'] == "Out of Home" and workBook[workSheetVertical]['F'+str(i)].value != 'N/A'):
                #     link = workBook[workSheetVertical]['F'+str(i)].value
                #     workBook[workSheetVertical]['F'+str(i)].value = 'Location'
                #     workBook[workSheetVertical]['F'+str(i)].hyperlink = link
                workBook[workSheetVertical]['B'+str(i)].fill = PatternFill("solid", fgColor="f7941d")
            if(workBook[workSheetVertical]['B'+str(i)].value):
                value = workBook[workSheetVertical]['B'+str(i)].value
                workBook[workSheetVertical]['B'+str(i)].value = 'View Image'
                workBook[workSheetVertical]['B'+str(i)].hyperlink = value
                # if(row['VerticalName'] == "Out of Home" and workBook[workSheetVertical]['F'+str(i)].value != 'N/A'):
                #     link = workBook[workSheetVertical]['F'+str(i)].value
                #     workBook[workSheetVertical]['F'+str(i)].value = 'Location'
                #     workBook[workSheetVertical]['F'+str(i)].hyperlink = link
                workBook[workSheetVertical]['B'+str(i)].fill = PatternFill("solid", fgColor="f7941d")
                i = i+1
        setSheet(workBook[workSheetVertical])
        for cell in workBook[workSheetVertical]['6']:
            if cell.value:
                cell.font = Font(bold=True, size=12, color="FFFFFFFF")
                cell.fill = PatternFill("solid", fgColor="f7941d")
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrapText=True)
        setAutoWidth(workBook[workSheetVertical])
        workBook[workSheetVertical].merge_cells('A3:D3')
        verticalNameCell = workBook[workSheetVertical].cell(row=3, column=1)
        verticalNameCell.alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True)
        verticalNameCell.font = Font(bold=True, size=14)
        verticalNameCell.value = row['VerticalName']+' Estimate'
    return workBook


def setAutoWidth(WorkSheet):
    for column_cells in WorkSheet.columns:
        new_column_length = max(len(str(cell.value))
                                for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            WorkSheet.column_dimensions[new_column_letter].width = 22
            WorkSheet.column_dimensions[new_column_letter].height = 100
            for cell in column_cells:
                if cell.value == "Availability":
                    WorkSheet.column_dimensions[new_column_letter].width = 40
                if cell.value == "Media Details":
                    WorkSheet.column_dimensions[new_column_letter].width = 100
                if cell.value == "Media Name":
                    WorkSheet.column_dimensions[new_column_letter].width = 50
                if cell.value == "Countries Covered":
                    WorkSheet.column_dimensions[new_column_letter].width = 40
                if cell.value == "View Image":
                    cell.font =  Font(bold=True, size=12, color="0056b3")
        for cell in column_cells:
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrapText=True)


def setSheet(WorkSheet):
    WorkSheet.sheet_view.showGridLines = False
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='000000'),
        right=Side(border_style=BORDER_THIN, color='000000'),
        top=Side(border_style=BORDER_THIN, color='000000'),
        bottom=Side(border_style=BORDER_THIN, color='000000')
    )

    no_fill = PatternFill(fill_type=None)
    no_border = Border(
        left=Side(border_style=None, color='FFFFFFFF'),
        right=Side(border_style=None, color='FFFFFFFF'),
        top=Side(border_style=None, color='FFFFFFFF'),
        bottom=Side(border_style=None, color='FFFFFFFF'),
    )
    for column_cells in WorkSheet.rows:
        for cell in column_cells:

            if cell.value or cell.value == 0:
                cell.font = Font(size=11)
                cell.fill = PatternFill("solid", fgColor="f3f3f3")
                cell.border = thin_border


@csrf_exempt
def downloadExcel(request, *args, **kwargs):
    excelData = JSONParser().parse(request)
    workBook = Workbook()

    DEFAULT_ROW_HEIGHT = 15
    overview = workBook.active
    overview.title = 'Overview'
    workBook = setVerticalPage(workBook, json.loads(excelData['Opportunity']))
    overview = setOverviewPage(overview, excelData)
    isExist = os.path.exists('media/estimate/excel/')
    if not isExist:
        os.makedirs('media/estimate/excel/')
    workBook.save('media/estimate/excel/'+excelData['PlanName']+".xlsx")
    # workBook = load_workbook('pylenin.xlsx')
    # print(workBook.sheetnames)

    # response = HttpResponse(workBook, content_type='application/xlsx')
    # filename = "Estimate_%s.xlsx" %(excelData['PlanName'])
    # content = "attachment; filename=%s" %(filename)
    # response['Content-Disposition'] = content
    # return response

    return JsonResponse('/media/estimate/excel/'+excelData['PlanName']+".xlsx", safe=False)
