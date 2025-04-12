from bs4 import BeautifulSoup
from django.db.models import Q
from django.db import connection
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from rest_framework.parsers import JSONParser
from django.http.response import JsonResponse
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from app_admin.models import AdFormat, AdUnitSpecificCategory, AdunitPosition, AdunitSize, AgeGroup, CityRegion, CountryEvent, EstimatedReach, GenderGroup, IncomeGroup, Language, LocationType, MatchingCategory, MediaCategory, NationalityCommunity, NearByAdvantage
from app_admin.serializers import AdFormatSerializer, AdUnitSpecificCategorySerializer, AdunitPositionSerializer, AdunitSizeSerializer, AgeGroupSerializer, CityRegionSerializer, CountryEventSerializer, EstimatedReachSerializer, GenderGroupSerializer, IncomeGroupSerializer, LanguageSerializer, LocationTypeSerializer, MatchingCategorySerializer, MediaCategorySerializer, NationalityCommunitySerializer, NearByAdvantageSerializer
from app_default.models import AdUnitCategory, Vertical
from app_default.serializers import AdUnitCategorySerializer, VerticalSerializer
from app_publisher.models import PublisherRegister
from app_publisher.serializers import PublisherRegisterSerializer
from app_vertical_media.models import Media, MediaAdType, MediaAdTypeImage

from app_vertical_media.serializers import FilterMediaSerializer, MediaAdTypeSerializer, MediaSerializer, SaveMediaAdTypeSerializer, SaveMediaSerializer, UploadMediaAdTypeImageSerializer, UploadMediaAdTypeSerializer, UploadMediaSerializer

# Create your views here.
import json
import gzip
from django.http import HttpResponse
from django.db.models import Max


# @csrf_exempt
# def RestMedia(request):
#     # try:
#     if request.method == 'GET':
#         media = Media.objects.all()
#         mediaSerializer = MediaSerializer(
#             media, many=True)

#         # for x in mediaSerializer.data:
#         #     try:
#         #         x['CountryEventJSON'] = json.loads(x['CountryEvent'])
#         #     except:
#         #         x['CountryEventJSON'] = x['CountryEvent']
#         #     try:
#         #         x['CityRegionJSON'] = json.loads(x['CityRegion'])
#         #     except:
#         #         x['CityRegionJSON'] = x['CityRegion']
#         #     try:
#         #         x['LanguageJSON'] = json.loads(x['Language'])
#         #     except:
#         #         x['LanguageJSON'] = x['Language']
#         #     try:
#         #         x['SEOMetaKeywordJSON'] = json.loads(x['SEOMetaKeyword'])
#         #     except:
#         #         x['SEOMetaKeywordJSON'] = x['SEOMetaKeyword']

#         #     media = Media.objects.get(
#         #         MediaId=x['MediaId'])
#         #     mediaSerializer = SaveMediaSerializer(
#         #         media, data=x)
#         #     if mediaSerializer.is_valid():
#         #         mediaSerializer.save()
#         return JsonResponse(mediaSerializer.data, safe=False)
#     else:
#         return JsonResponse("Invalid request", safe=False)
#     # except:
#     #     return JsonResponse("Invalid payload", safe=False)


# @csrf_exempt
# def RestMediaAdType(request):
#     # try:
#     if request.method == 'GET':
#         media = MediaAdType.objects.filter(
#             Q(AvailabilityJSON__isnull=True) |
#             Q(NearByAdvantagesJSON__isnull=True) |
#             Q(MatchingCategoriesJSON__isnull=True) |
#             Q(AgeGroupJSON__isnull=True) |
#             Q(GenderGroupJSON__isnull=True) |
#             Q(IncomeGroupJSON__isnull=True)|
#             Q(AvailabilityJSON="") |
#             Q(NearByAdvantagesJSON="") |
#             Q(MatchingCategoriesJSON="") |
#             Q(AgeGroupJSON="") |
#             Q(GenderGroupJSON="") |
#             Q(IncomeGroupJSON=""))
#         mediaSerializer = MediaAdTypeSerializer(
#             media, many=True)

#         for x in mediaSerializer.data:
#             try:
#                 x['AvailabilityJSON'] = json.loads(x['Availability'])
#             except:
#                 x['AvailabilityJSON'] = x['Availability']
#             try:
#                 x['NearByAdvantagesJSON'] = json.loads(x['NearByAdvantages'])
#             except:
#                 x['NearByAdvantagesJSON'] = x['NearByAdvantages']
#             try:
#                 x['MatchingCategoriesJSON'] = json.loads(
#                     x['MatchingCategories'])
#             except:
#                 x['MatchingCategoriesJSON'] = x['MatchingCategories']
#             try:
#                 x['AgeGroupJSON'] = json.loads(x['AgeGroup'])
#             except:
#                 x['AgeGroupJSON'] = x['AgeGroup']
#             try:
#                 x['GenderGroupJSON'] = json.loads(x['GenderGroup'])
#             except:
#                 x['GenderGroupJSON'] = x['GenderGroup']
#             try:
#                 x['IncomeGroupJSON'] = json.loads(x['IncomeGroup'])
#             except:
#                 x['IncomeGroupJSON'] = x['IncomeGroup']

#             media = MediaAdType.objects.get(
#                 MediaAdTypeId=x['MediaAdTypeId'])
#             mediaSerializer = SaveMediaAdTypeSerializer(
#                 media, data=x)
#             if mediaSerializer.is_valid():
#                 mediaSerializer.save()
#         return JsonResponse("done", safe=False)
#     else:
#         return JsonResponse("Invalid request", safe=False)
#     # except:
#     #     return JsonResponse("Invalid payload", safe=False)


# ["8db1f930-8f0f-4eb4-8638-5915c016aa9f", "48a0c457-ec86-4dd3-ba2c-f807b80cd1ad", "0c476141-30bd-4ad6-a8c6-92abdf2cc31a",
#     "a8a9f316-59b8-41e2-ba09-472bbfed1831", "2a17af96-1831-4b04-8bd0-60de8dedf2c3"]
# ["8db1f930-8f0f-4eb4-8638-5915c016aa9f", "48a0c457-ec86-4dd3-ba2c-f807b80cd1ad", "0c476141-30bd-4ad6-a8c6-92abdf2cc31a",
#     "a8a9f316-59b8-41e2-ba09-472bbfed1831", "2a17af96-1831-4b04-8bd0-60de8dedf2c3"]
import re
# as per recommendation from @freylis, compile once only
CLEANR = re.compile('<.*?>')


@csrf_exempt
def RestMedia(request):
    if request.method == 'GET':
        media = Media.objects.filter(
            Q(BenefitsOfAdvertisingNonHTML__isnull=True))
        mediaSerializer = MediaSerializer(
            media, many=True)
        for x in mediaSerializer.data:
            x['BenefitsOfAdvertisingNonHTML'] = ""
            if x['BenefitsOfAdvertisingNonHTML']:
                x['BenefitsOfAdvertisingNonHTML'] = cleanhtml(
                    x['BenefitsOfAdvertising'])
            media = Media.objects.get(
                MediaId=x['MediaId'])
            mediaSerializer = SaveMediaSerializer(
                media, data=x)
            if mediaSerializer.is_valid():
                mediaSerializer.save()
        return JsonResponse("Done", safe=False)


@csrf_exempt
def RestMediaAdType(request):
    if request.method == 'GET':
        media = MediaAdType.objects.filter(
            Q(MediaAdUnitDescriptionNonHTML__isnull=True) |
            Q(TermsAndConditionsNonHTML__isnull=True))
        mediaSerializer = MediaAdTypeSerializer(
            media, many=True)
        for x in mediaSerializer.data:
            x['TermsAndConditionsNonHTML'] = ""
            x['MediaAdUnitDescriptionNonHTML'] = ""
            if x['TermsAndConditions']:
                x['TermsAndConditionsNonHTML'] = cleanhtml(
                    x['TermsAndConditions'])
            if x['MediaAdUnitDescription']:
                x['MediaAdUnitDescriptionNonHTML'] = cleanhtml(
                    x['MediaAdUnitDescription'])
            media = MediaAdType.objects.get(
                MediaAdTypeId=x['MediaAdTypeId'])
            mediaSerializer = SaveMediaAdTypeSerializer(
                media, data=x)
            if mediaSerializer.is_valid():
                mediaSerializer.save()
        return JsonResponse("Done", safe=False)


def cleanhtml(raw_html):
    cleantext = re.sub(CLEANR, '', raw_html)
    return cleantext


@csrf_exempt
def CURDMediaLogo(request,  pk=""):
    if request.method == 'DELETE':
        media = Media.objects.get(MediaId=pk)
        media.MediaImage.delete(save=True)
        return JsonResponse("Deleted successfully", safe=False)


@csrf_exempt
def CURDMediaKit(request,  pk=""):
    if request.method == 'DELETE':
        media = Media.objects.get(MediaId=pk)
        media.MediaKit.delete(save=True)
        return JsonResponse("Deleted successfully", safe=False)


@csrf_exempt
def AllMediaApi(request):
    try:
        if request.method == 'GET':
            media = Media.objects.all()
            mediaSerializer = FilterMediaSerializer(
                media, many=True)
            # return JsonResponse(mediaSerializer.data, safe=False)
            json_data = json.dumps(mediaSerializer.data).encode('utf-8')

            compressed_data = gzip.compress(json_data)

            response = HttpResponse(
                compressed_data, content_type='application/json')
            response['Content-Encoding'] = 'gzip'
            return response
        else:
            return JsonResponse("Invalid request", safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)

#


@csrf_exempt
def MediaListingApi(request, VerticalId="", CountryId="", Limit="0", offset="30"):
    # try:
    if request.method == 'GET':
        cursor = connection.cursor()
        cursor.execute("SELECT MediaId FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND VerticalType='" +
                       VerticalId + "' AND  CountryEvent  LIKE '%%" + CountryId + "%%' LIMIT "+Limit+", "+offset+"")

        Count = connection.cursor()
        Count.execute("SELECT COUNT(MediaId) as Count FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND VerticalType='" +
                      VerticalId + "' AND  CountryEvent  LIKE '%%" + CountryId + "%%'")

        MediaId = "''"
        for x in cursor.fetchall():
            MediaId = MediaId+",'"+x[0]+"'"
        media = Media.objects.raw(
            "SELECT * FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND MediaId IN (" + MediaId + ")")

        mediaSerializer = MediaSerializer(
            media, many=True)

        mediaAdType = MediaAdType.objects.raw("SELECT * FROM app_vertical_media_mediaadtype WHERE MediaId IN (SELECT MediaId FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND VerticalType='" +
                                              VerticalId + "' AND  CountryEvent  LIKE '%%" + CountryId + "%%' LIMIT "+Limit+", "+offset+")")

        mediaAdTypeSerializer = MediaAdTypeSerializer(
            mediaAdType, many=True)

        mediaAdType = MediaAdType.objects.raw(
            "SELECT * FROM app_vertical_media_mediaadtype WHERE MediaId IN (" + MediaId + ")")

        mediaAdTypeSerializer = MediaAdTypeSerializer(
            mediaAdType, many=True)

        plan = MediaAdTypeImage.objects.raw(
            "SELECT * FROM app_vertical_media_mediaadtypeimage WHERE MediaAdTypeId IN (SELECT MediaAdTypeId FROM app_vertical_media_mediaadtype WHERE MediaId IN (" + MediaId + "))")
        mediaAdTypeImageSerializer = UploadMediaAdTypeImageSerializer(
            plan, many=True)

        Data = []
        Data.append(mediaSerializer.data)
        Data.append(mediaAdTypeSerializer.data)
        Data.append(mediaAdTypeImageSerializer.data)
        Data.append(Count.fetchall())
        return JsonResponse(Data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def MediaListingAudienceApi(request, AudienceId="", CountryId="", Limit="0", offset="30"):
    # try:
    if request.method == 'GET':
        cursor = connection.cursor()

        aduintsubquery = "SELECT MediaId FROM app_vertical_media_mediaadtype WHERE Status IN ('2','3') AND ShowOnOff='1' AND MatchingCategories LIKE '%%" + \
            AudienceId + "%%'"

        cursor.execute("SELECT MediaId FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND MediaId IN (" + aduintsubquery + ")  AND  CountryEvent  LIKE '%%" +
                       CountryId + "%%' LIMIT "+Limit+", "+offset+"")

        Count = connection.cursor()
        Count.execute("SELECT COUNT(MediaId) as Count FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND MediaId IN (" +
                      aduintsubquery + ")  AND  CountryEvent  LIKE '%%" + CountryId + "%%'")

        MediaId = "''"
        for x in cursor.fetchall():
            MediaId = MediaId+",'"+x[0]+"'"
        media = Media.objects.raw(
            "SELECT * FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND MediaId IN (" + MediaId + ")")

        mediaSerializer = MediaSerializer(
            media, many=True)

        mediaAdType = MediaAdType.objects.raw(
            "SELECT * FROM app_vertical_media_mediaadtype WHERE MediaId IN (SELECT MediaId FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND MediaId IN (" + aduintsubquery + ")  AND  CountryEvent  LIKE '%%" + CountryId + "%%' LIMIT "+Limit+", "+offset+")")

        mediaAdTypeSerializer = MediaAdTypeSerializer(
            mediaAdType, many=True)

        mediaAdType = MediaAdType.objects.raw(
            "SELECT * FROM app_vertical_media_mediaadtype WHERE MediaId IN (" + MediaId + ")")

        mediaAdTypeSerializer = MediaAdTypeSerializer(
            mediaAdType, many=True)

        plan = MediaAdTypeImage.objects.raw(
            "SELECT * FROM app_vertical_media_mediaadtypeimage WHERE MediaAdTypeId IN (SELECT MediaAdTypeId FROM app_vertical_media_mediaadtype WHERE MediaId IN (" + MediaId + "))")
        mediaAdTypeImageSerializer = UploadMediaAdTypeImageSerializer(
            plan, many=True)

        Data = []
        Data.append(mediaSerializer.data)
        Data.append(mediaAdTypeSerializer.data)
        Data.append(mediaAdTypeImageSerializer.data)
        Data.append(Count.fetchall())
        return JsonResponse(Data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def MediaForSpecificCategoryApi(request):
    # try:
    #     'VerticalNSpecificCategoryId': this.VerticalNSpecificCategoryId,
    #   'VerticalNSpecificCategoryName': this.VerticalNSpecificCategoryName,
    #   'VerticalNSpecificCategoryType': this.VerticalNSpecificCategoryType,
    #   'Country': this.filterForm.value.Country,
    #   'AudienceId': this.AudienceId,
    #   '': this.Offset,
    #   '': this.Limit
    mediaData = JSONParser().parse(request)
    AudienceId = mediaData['AudienceId']
    CountryId = mediaData['Country']
    Limit = mediaData['Limit']
    offset = mediaData['Offset']
    MediaCount = mediaData['MediaCount']
    City = mediaData['City']
    Language = mediaData['Language']
    CampaignType = mediaData['CampaignType']
    if request.method == 'POST':
        Data = []
        if mediaData['VerticalNSpecificCategoryType'] == 'Vertical':
            VerticalId = mediaData['VerticalNSpecificCategoryId']
            query = ""
            if CountryId:
                query = query + \
                    " AND JSON_CONTAINS(CountryEventJSON, '" + \
                    '"' + CountryId+'"'+"') "
            if VerticalId:
                query = query+" AND VerticalType='" + VerticalId + "' "
            if City:
                query = query + \
                    " AND JSON_CONTAINS(CityRegionJSON, '"+'"' + City+'"'+"') "
            if Language:
                query = query + \
                    " AND JSON_CONTAINS(LanguageJSON, '" + \
                    '"' + Language+'"'+"') "
            if MediaCount:
                Count = connection.cursor()
                Count.execute("SELECT COUNT(MediaId) as Count FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND VerticalType='" +
                              VerticalId + "' AND  CountryEvent  LIKE '%%" + CountryId + "%%'")
                MediaCount = Count.fetchall()[0][0]

            # media = Media.objects.raw(
            #     "SELECT * FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND VerticalType='" +
            #     VerticalId + "' AND  CountryEvent  LIKE '%%" + CountryId + "%%'  LIMIT "+offset+", "+Limit+"")
            media = Media.objects.raw(
                "SELECT * FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' " + query + "  LIMIT "+offset+", "+Limit+"")

            mediaSerializer = MediaSerializer(
                media, many=True)

            MediaId = "''"
            for x in mediaSerializer.data:
                MediaId = MediaId+",'"+x['MediaId']+"'"

            mediaAdType = MediaAdType.objects.raw(
                "SELECT * FROM app_vertical_media_mediaadtype WHERE MediaId IN (" + MediaId + ")")

            mediaAdTypeSerializer = MediaAdTypeSerializer(
                mediaAdType, many=True)

            Data.append(mediaSerializer.data)
            Data.append(mediaAdTypeSerializer.data)
            Data.append(MediaCount)
        elif mediaData['VerticalNSpecificCategoryType'] == 'SpecificCategory':
            SpecificCategoryId = mediaData['VerticalNSpecificCategoryId']
            if MediaCount:
                Count = connection.cursor()
                Count.execute("SELECT COUNT(MediaId) as Count FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND MediaId IN (SELECT MediaId FROM app_vertical_media_mediaadtype WHERE SpecificCategory='" +
                              SpecificCategoryId+"') AND  CountryEvent  LIKE '%%" + CountryId + "%%'")
                MediaCount = Count.fetchall()[0][0]

            media = Media.objects.raw(
                "SELECT * FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' AND  MediaId IN (SELECT MediaId FROM app_vertical_media_mediaadtype WHERE SpecificCategory='"+SpecificCategoryId+"') AND  CountryEvent  LIKE '%%" + CountryId + "%%' LIMIT "+offset+", "+Limit+"")

            mediaSerializer = MediaSerializer(
                media, many=True)

            MediaId = "''"
            for x in mediaSerializer.data:
                MediaId = MediaId+",'"+x['MediaId']+"'"

            mediaAdType = MediaAdType.objects.raw(
                "SELECT * FROM app_vertical_media_mediaadtype WHERE  SpecificCategory='"+SpecificCategoryId+"' ")

            mediaAdTypeSerializer = MediaAdTypeSerializer(
                mediaAdType, many=True)

            Data.append(mediaSerializer.data)
            Data.append(mediaAdTypeSerializer.data)
            Data.append(MediaCount)
        return JsonResponse(Data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def PublisherMediaListingApi(request):
    mediaData = JSONParser().parse(request)

    try:
        media = Media.objects.filter(
            VerticalType=mediaData['VerticalId'], PublisherId=mediaData['PublisherId'])
        mediaSerializer = MediaSerializer(
            media, many=True)

        Data = []
        for x in mediaSerializer.data:
            media = []

            Status0 = connection.cursor()
            Status1 = connection.cursor()
            Status2 = connection.cursor()
            Status3 = connection.cursor()
            Status4 = connection.cursor()
            Status5 = connection.cursor()
            Status6 = connection.cursor()
            Status7 = connection.cursor()

            Status0.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='0'")
            Status1.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='1'")
            Status2.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='2'")
            Status3.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='3'")
            Status4.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='4'")
            Status5.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='5'")
            Status6.execute(
                "SELECT COUNT(ShowOnOff) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND ShowOnOff='0'")
            Status7.execute(
                "SELECT COUNT(ShowOnOff) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND ShowOnOff='1'")

            StatusCount = []
            StatusCount.append(Status0.fetchall()[0][0])
            StatusCount.append(Status1.fetchall()[0][0])
            StatusCount.append(Status2.fetchall()[0][0])
            StatusCount.append(Status3.fetchall()[0][0])
            StatusCount.append(Status4.fetchall()[0][0])
            StatusCount.append(Status5.fetchall()[0][0])
            StatusCount.append(Status6.fetchall()[0][0])
            StatusCount.append(Status7.fetchall()[0][0])
            # publisher=[];
            # if x['PublisherId']:
            #     profile = PublisherRegister.objects.get(
            #         PublisherId=x['PublisherId'])
            #     profileSerializer = PublisherSerializer(
            #         profile)
            #     publisher = profileSerializer.data
            media.append(x)
            media.append(StatusCount)
            # media.append(publisher)
            Data.append(media)
        return JsonResponse(Data, safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def AdminMediaListingApi(request, VerticalId=""):

    try:
        media = Media.objects.exclude(Status=0).filter(
            VerticalType=VerticalId)
        mediaSerializer = MediaSerializer(
            media, many=True)

        Data = []
        for x in mediaSerializer.data:
            media = []

            Status0 = connection.cursor()
            Status1 = connection.cursor()
            Status2 = connection.cursor()
            Status3 = connection.cursor()
            Status4 = connection.cursor()
            Status5 = connection.cursor()
            Status6 = connection.cursor()
            Status7 = connection.cursor()

            Status0.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='0'")
            Status1.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='1'")
            Status2.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='2'")
            Status3.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='3'")
            Status4.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='4'")
            Status5.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='5'")
            Status6.execute(
                "SELECT COUNT(ShowOnOff) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND ShowOnOff='0'")
            Status7.execute(
                "SELECT COUNT(ShowOnOff) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND ShowOnOff='1'")

            StatusCount = []
            StatusCount.append(0)
            StatusCount.append(Status1.fetchall()[0][0])
            StatusCount.append(Status2.fetchall()[0][0])
            StatusCount.append(Status3.fetchall()[0][0])
            StatusCount.append(Status4.fetchall()[0][0])
            StatusCount.append(Status5.fetchall()[0][0])
            StatusCount.append(Status6.fetchall()[0][0])
            StatusCount.append(Status7.fetchall()[0][0])
            media.append(x)
            media.append(StatusCount)
            Data.append(media)
        return JsonResponse(Data, safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def ManageMediaListingApi(request, Limit="0", offset="30"):
    # try:
    if request.method == 'POST':
        mediaData = JSONParser().parse(request)
        cursor = connection.cursor()
        PublisherSQL = ""
        mediacategorySQL = ""
        if mediaData['MediaCategory']:
            mediacategorySQL = " AND   MediaCategory='" + \
                mediaData['MediaCategory'] + "' "
        if mediaData['PublisherId']:
            PublisherSQL = " AND   PublisherId='" + \
                mediaData['PublisherId'] + "' "
        # else:
        #     mediacategorySQL="AND   MediaCategory IS NULL "
        countrySQL = ""
        if mediaData['Country']:
            countrySQL = "AND  CountryEvent  LIKE '%%" + \
                mediaData['Country'] + "%%' "

        cursor.execute(
            "SELECT MediaId FROM app_vertical_media_media WHERE  VerticalType='" +
            mediaData['VerticalId'] + "' " + countrySQL + mediacategorySQL + PublisherSQL+" ORDER BY SortId DESC  LIMIT "+Limit+", "+offset+"")

        Count = connection.cursor()
        Count.execute("SELECT COUNT(MediaId) as Count FROM app_vertical_media_media WHERE  VerticalType='" +
                      mediaData['VerticalId'] + "' " + countrySQL+mediacategorySQL + PublisherSQL+" ORDER BY SortId DESC")

        MediaId = "''"
        for x in cursor.fetchall():
            MediaId = MediaId+",'"+x[0]+"'"
        media = Media.objects.raw(
            "SELECT * FROM app_vertical_media_media WHERE MediaId IN (" + MediaId + ")  ORDER BY SortId DESC")

        mediaSerializer = MediaSerializer(
            media, many=True)

        Data = []
        for x in mediaSerializer.data:
            media = []

            Status0 = connection.cursor()
            Status1 = connection.cursor()
            Status2 = connection.cursor()
            Status3 = connection.cursor()
            Status4 = connection.cursor()
            Status5 = connection.cursor()
            Status6 = connection.cursor()
            Status7 = connection.cursor()

            Status0.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='0'")
            Status1.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='1'")
            Status2.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='2'")
            Status3.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='3'")
            Status4.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='4'")
            Status5.execute(
                "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND Status='5'")
            Status6.execute(
                "SELECT COUNT(ShowOnOff) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND ShowOnOff='0'")
            Status7.execute(
                "SELECT COUNT(ShowOnOff) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + x['MediaId'] + "' AND ShowOnOff='1'")

            StatusCount = []
            StatusCount.append(0)
            StatusCount.append(Status1.fetchall()[0][0])
            StatusCount.append(Status2.fetchall()[0][0])
            StatusCount.append(Status3.fetchall()[0][0])
            StatusCount.append(Status4.fetchall()[0][0])
            StatusCount.append(Status5.fetchall()[0][0])
            StatusCount.append(Status6.fetchall()[0][0])
            StatusCount.append(Status7.fetchall()[0][0])
            media.append(x)
            media.append(StatusCount)
            Data.append(media)
        FinalData = []
        FinalData.append(Data)
        FinalData.append(Count.fetchall())
        return JsonResponse(FinalData, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def ManageMediaListingApi2(request, Limit="0", offset="30"):
    # try:
    if request.method == 'POST':
        mediaData = JSONParser().parse(request)
        cursor = connection.cursor()
        PublisherSQL = ""
        mediacategorySQL = ""
        if mediaData['MediaCategory']:
            mediacategorySQL = " AND   MediaCategory='" + \
                mediaData['MediaCategory'] + "' "
        if mediaData['PublisherId']:
            PublisherSQL = " AND   PublisherId='" + \
                mediaData['PublisherId'] + "' "
        # else:
        #     mediacategorySQL="AND   MediaCategory IS NULL "
        countrySQL = ""
        if mediaData['Country']:
            countrySQL = "AND  CountryEvent  LIKE '%%" + \
                mediaData['Country'] + "%%' "

        # cursor.execute(
        #     "SELECT MediaId FROM app_vertical_media_media WHERE  VerticalType='" +
        #     mediaData['VerticalId'] + "' " + countrySQL + mediacategorySQL + PublisherSQL+" ORDER BY SortId DESC  LIMIT "+Limit+", "+offset+"")

        Count = connection.cursor()
        Count.execute("SELECT COUNT(MediaId) as Count FROM app_vertical_media_media WHERE  VerticalType='" +
                      mediaData['VerticalId'] + "' " + countrySQL+mediacategorySQL + PublisherSQL+" ORDER BY SortId DESC")

        # MediaId = "''"
        # for x in cursor.fetchall():
        #     MediaId = MediaId+",'"+x[0]+"'"
        media = Media.objects.raw(
            "SELECT * FROM app_vertical_media_media WHERE  VerticalType='" +
            mediaData['VerticalId'] + "' " + countrySQL + mediacategorySQL + PublisherSQL+" ORDER BY SortId DESC  LIMIT "+Limit+", "+offset+"")

        mediaSerializer = MediaSerializer(
            media, many=True)

        Data = []

        # Fetch status counts in a single query
        status_query = f"""
            SELECT MediaId,
                SUM(CASE WHEN Status = '0' THEN 1 ELSE 0 END) as Status0,
                SUM(CASE WHEN Status = '1' THEN 1 ELSE 0 END) as Status1,
                SUM(CASE WHEN Status = '2' THEN 1 ELSE 0 END) as Status2,
                SUM(CASE WHEN Status = '3' THEN 1 ELSE 0 END) as Status3,
                SUM(CASE WHEN Status = '4' THEN 1 ELSE 0 END) as Status4,
                SUM(CASE WHEN Status = '5' THEN 1 ELSE 0 END) as Status5,
                SUM(CASE WHEN ShowOnOff = '0' THEN 1 ELSE 0 END) as ShowOnOff0,
                SUM(CASE WHEN ShowOnOff = '1' THEN 1 ELSE 0 END) as ShowOnOff1
            FROM app_vertical_media_mediaadtype
            WHERE MediaId IN %s
            GROUP BY MediaId
        """

        media_ids = [row['MediaId'] for row in mediaSerializer.data]
        cursor.execute(status_query, [tuple(media_ids)])
        status_counts = {row[0]: [int(value) for value in row[1:]]
                         for row in cursor.fetchall()}

        data = []
        for media_item in mediaSerializer.data:
            media_id = media_item['MediaId']
            status_count = status_counts.get(media_id, [0] * 8)
            data.append([media_item, status_count])

        final_data = [data, Count.fetchall()]
        return JsonResponse(final_data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDMediaDetailsApi(request, MediaId=""):
    try:
        if request.method == 'GET':
            media = Media.objects.get(
                MediaId=MediaId)
            mediaSerializer = MediaSerializer(
                media)
            return JsonResponse(mediaSerializer.data, safe=False)
        else:
            return JsonResponse("Invalid request", safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def MediaStatusCountByIdApi(request, MediaId=""):
    # try:
    if MediaId:
        Status1 = connection.cursor()
        Status2 = connection.cursor()
        Status3 = connection.cursor()
        Status4 = connection.cursor()
        Status5 = connection.cursor()
        Status1.execute(
            "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + MediaId + "' AND Status='1'")
        Status2.execute(
            "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + MediaId + "' AND Status='2'")
        Status3.execute(
            "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + MediaId + "' AND Status='3'")
        Status4.execute(
            "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + MediaId + "' AND Status='4'")
        Status5.execute(
            "SELECT COUNT(Status) as count FROM app_vertical_media_mediaadtype WHERE MediaId='" + MediaId + "' AND Status='5'")

        Data = []
        Data.append('0')
        Data.append(Status1.fetchall()[0][0])
        Data.append(Status2.fetchall()[0][0])
        Data.append(Status3.fetchall()[0][0])
        Data.append(Status4.fetchall()[0][0])
        Data.append(Status5.fetchall()[0][0])
        return JsonResponse(Data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)

def saveToRagSystemByMediaId(MediaId=""):
    media = Media.objects.get(
        MediaId=MediaId)
    mediaSerializer = MediaSerializer(
        media)
    mediaAdType = MediaAdType.objects.filter(
        MediaId=MediaId)
    mediaAdTypeSerializer = MediaAdTypeSerializer(
        mediaAdType,many=True)
    data=[]
    for adunit in mediaAdTypeSerializer.data:
        data.append({
            'status':saveToRagSystemByAdUnitId(adunit['MediaAdTypeId'],adunit['MediaId']),
            'MediaAdTypeId':adunit['MediaAdTypeId'],
            'MediaId':adunit['MediaId'],
            })
    return data
import requests
from datetime import datetime
def saveToRagSystemByAdUnitId(AdUnitId="",MediaId=""):
    
    def get_units(vertical_name, media_ad_type):
        if media_ad_type == 'Package':
            return 'Package'
        if vertical_name == 'Out of Home':
            return 'Monthly'
        if vertical_name == 'Radio':
            return 'Spot'
        if vertical_name == 'Digital':
            return 'CPM'
        if vertical_name == 'Newspaper':
            return 'Ad Rate'
        if vertical_name == 'Magazine':
            return 'Ad Rate'
        if vertical_name == 'Television':
            return 'Spot'
        if vertical_name == 'Media Service':
            return 'Service'
        return ''  # Default empty if no condition is met
    media_mediaadtype = connection.cursor()
    sql="SELECT app_vertical_media_mediaadtype.SortId, app_vertical_media_mediaadtype.MediaId, VerticalType, MediaName, Category, PrimaryLanguage, TimeZone, MediaManagedBy, WebsiteLink, EventVenue, PublisherId, CountryEvent, CityRegion, Language, EstimateReached, BenefitsOfAdvertising, app_vertical_media_mediaadtype.PromoVideoLink, MediaKit, MapLink, Latitude, Longitude, IsHyperlocal, app_vertical_media_mediaadtype.Availability, MediaImage, app_vertical_media_mediaadtype.Status, ViewsCount, ModifiedDate, app_vertical_media_mediaadtype.ShowOnOff, WebStatsOnOff, RecommendedMedia, MediaCategory, ViewOnlyMediaStore, ContactPerson, app_vertical_media_mediaadtype.Identification, SEOMetaDescription, SEOMetaKeyword, SEOMetaTitle, SEOMetaKeywordJSON, BenefitsOfAdvertisingNonHTML, MediaAdTypeId, Size, OutdoorSize, CreativeSize, AdFormat, NumberOfSpots, PromoAudio, ProgramStartTime, ProgramEndTime, OrderUnit, UnitType, OrderingQuantity, CreativesRequired, Position, DigitalPosition, MediaAdType, SiteName, LocationType, VerticalTypeId, NoOfScreen, MediaAdUnitDescription, IsDigital, SingleRotationTime, DurationOfCreative, PackageName, PackageType, StartDate, EndDate, NoOfDays, MonthlyRentalRate, Discount, NetCostBeforeTax, ProductionCost, Tax, NetCost, RateValidaty, AdUnitCategory, AdTypeMediaKit, AdTypeEstimateReached, NearByAdvantages, MatchingCategories, AgeGroup, GenderGroup, IncomeGroup, NationalityCommunity, RunAsOffer, IsLocalTaxApplied, IsNeedApproval, NoOfDaysRequired, MediaAdTypeImage, TermsAndConditions, ThumbnailImage, Currency, RateDisplay, SpecificCategory, CostPerThousand, CostPerThousandDisplay, AgeGroupJSON, AvailabilityJSON, GenderGroupJSON, IncomeGroupJSON, MatchingCategoriesJSON, NearByAdvantagesJSON, MediaAdUnitDescriptionNonHTML, TermsAndConditionsNonHTML, PriceRangeNote from app_vertical_media_media JOIN app_vertical_media_mediaadtype ON app_vertical_media_media.MediaId=app_vertical_media_mediaadtype.MediaId AND app_vertical_media_mediaadtype.MediaAdTypeId='"+AdUnitId+"';"
    media_mediaadtype.execute(sql)
    field_names = [i[0] for i in media_mediaadtype.description]
    mediaData = []
    for media in media_mediaadtype.fetchall():
        j = 0
        media_data = {}
        for col in media:
            if j < len(field_names):
                media_data[str(field_names[j])] = col
                j = j + 1
        mediaData.append(media_data)
    ragMedaiData=mediaData[0]
    verticalCategory = Vertical.objects.get(
        VerticalId=ragMedaiData['VerticalType'])
    verticalCategorySerializer = VerticalSerializer(
        verticalCategory)
    city_regions = {c['CityId']: c['CityRegionName'] for c in CityRegionSerializer(CityRegion.objects.all(), many=True).data}
    countries = {c['CountryEventId']: c['CountryEventName'] for c in CountryEventSerializer(CountryEvent.objects.all(), many=True).data}
    incomeGroup = {c['IncomeGroupId']: c['IncomeGroup'] for c in IncomeGroupSerializer(IncomeGroup.objects.all(), many=True).data}
    city_region_data = [city_regions.get(id, "") for id in json.loads(ragMedaiData['CityRegion'])]
    incomeGroup_data = [incomeGroup.get(id, "") for id in json.loads(ragMedaiData['IncomeGroup'])]
    country_data = [countries.get(id, "") for id in json.loads(ragMedaiData['CountryEvent'])]
                
    ragMedaiData['VerticalName'] = verticalCategorySerializer.data['VerticalName']
    ragMedaiData['Units']=get_units(ragMedaiData['VerticalName'], ragMedaiData['MediaAdType'])
    ragMedaiData['City']=city_region_data
    # ragMedaiData['Country']=country_data
    ragMedaiData['IncomeGroup']=incomeGroup_data
    # Convert datetime objects to string format (ISO 8601)
    for key in ragMedaiData:
        if isinstance(ragMedaiData[key], datetime):
            ragMedaiData[key] = ragMedaiData[key].isoformat()

    payload={
        "collection_name_list": country_data,
        "media_category":ragMedaiData['VerticalName'],
        "payload": ragMedaiData
    }
    url="https://ai.globalmediakit.com/chat/update-collection/"
    method="PUT"
    response = requests.put(url, json=payload)
    return {'status':response.status_code,'payload':payload}

@csrf_exempt
def CURDMediaApi(request, VerticalId="", pk=""):
    try:
        if request.method == 'GET':
            if not VerticalId:
                media = Media.objects.filter(
                    ShowOnOff=1)
                mediaSerializer = FilterMediaSerializer(
                    media, many=True)
                return JsonResponse(mediaSerializer.data, safe=False)
            elif not pk:
                media = Media.objects.filter(
                    VerticalType=VerticalId)
                mediaSerializer = MediaSerializer(
                    media, many=True)
                return JsonResponse(mediaSerializer.data, safe=False)
            else:
                media = Media.objects.get(
                    MediaId=pk)
                mediaSerializer = MediaSerializer(
                    media)
                return JsonResponse(mediaSerializer.data, safe=False)
        elif request.method == 'POST':
            mediaData = JSONParser().parse(request)

            try:
                mediaData['CountryEventJSON'] = json.loads(
                    mediaData['CountryEvent'])
            except:
                mediaData['CountryEventJSON'] = []
            try:
                mediaData['CityRegionJSON'] = json.loads(mediaData['CityRegion'])
            except:
                mediaData['CityRegionJSON'] = []
            try:
                mediaData['LanguageJSON'] = json.loads(mediaData['Language'])
            except:
                mediaData['LanguageJSON'] = []
            try:
                mediaData['BenefitsOfAdvertisingNonHTML'] = cleanhtml(
                mediaData['BenefitsOfAdvertising'])
            except:
                mediaData['BenefitsOfAdvertisingNonHTML'] = ""
            try:
                mediaData['SEOMetaKeywordJSON'] = json.loads(
                mediaData['SEOMetaKeyword'])
            except:
                mediaData['SEOMetaKeywordJSON'] = []
            mediaSerializer = SaveMediaSerializer(
                data=mediaData)
            if mediaSerializer.is_valid():
                mediaSerializer.save()
                mediaAd = mediaSerializer.data
                last_id = Media.objects.aggregate(
                    Max('Identification'))['Identification__max']
                mediaAd['Identification'] = last_id+1
                mediaAdType = Media.objects.get(
                    MediaId=mediaSerializer.data['MediaId'])
                mediaSerializer2 = SaveMediaSerializer(
                    mediaAdType, mediaAd)
                if mediaSerializer2.is_valid():
                    mediaSerializer2.save()
                    return JsonResponse(mediaSerializer.data['MediaId'], safe=False)
            return JsonResponse("Failed", safe=False)
        elif request.method == 'PUT':
            mediaData = JSONParser().parse(request)
            try:
                mediaData['CountryEventJSON'] = json.loads(
                    mediaData['CountryEvent'])
            except:
                mediaData['CountryEventJSON'] = []
            try:
                mediaData['CityRegionJSON'] = json.loads(mediaData['CityRegion'])
            except:
                mediaData['CityRegionJSON'] = []
            try:
                mediaData['LanguageJSON'] = json.loads(mediaData['Language'])
            except:
                mediaData['LanguageJSON'] = []
            try:
                mediaData['BenefitsOfAdvertisingNonHTML'] = cleanhtml(
                mediaData['BenefitsOfAdvertising'])
            except:
                mediaData['BenefitsOfAdvertisingNonHTML'] = ""
            try:
                mediaData['SEOMetaKeywordJSON'] = json.loads(
                mediaData['SEOMetaKeyword'])
            except:
                mediaData['SEOMetaKeywordJSON'] = []
            media = Media.objects.get(
                MediaId=mediaData['MediaId'])
            mediaSerializer = SaveMediaSerializer(
                media, data=mediaData)
            if mediaSerializer.is_valid():
                mediaSerializer.save()
                saveToRagSystemByMediaId(mediaSerializer.data['MediaId'])
                return JsonResponse("Update successfully", safe=False)
            return JsonResponse("Failed to Update", safe=False)
        elif request.method == 'DELETE':
            media = Media.objects.get(MediaId=pk)
            media.MediaKit.delete(save=True)
            media.MediaImage.delete(save=True)
            media.delete()
            return JsonResponse("Deleted successfully", safe=False)
        else:
            return JsonResponse("Invalid request", safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDPublisherMedia(request, PublisherId="", VerticalId=""):
    # try:
    if VerticalId:
        media = Media.objects.filter(
            VerticalType=VerticalId, PublisherId=PublisherId)
        mediaSerializer = MediaSerializer(
            media, many=True)
        return JsonResponse(mediaSerializer.data, safe=False)
    else:
        # media = Media.objects.filter(
        #     PublisherId=PublisherId)
        # mediaSerializer = MediaSerializer(
        #     media, many=True)
        # return JsonResponse(mediaSerializer.data, safe=False)

        cursor = connection.cursor()
        cursor.execute(
            "SELECT MediaId FROM app_vertical_media_media WHERE Status IN ('2','3') AND ShowOnOff='1' AND PublisherId='" + PublisherId + "'")

        Count = connection.cursor()
        Count.execute(
            "SELECT COUNT(MediaId) as Count FROM app_vertical_media_media WHERE Status IN ('2','3') AND ShowOnOff='1' AND PublisherId='" + PublisherId + "'")

        MediaId = "''"
        for x in cursor.fetchall():
            MediaId = MediaId+",'"+x[0]+"'"
        media = Media.objects.raw(
            "SELECT * FROM app_vertical_media_media WHERE MediaId IN (" + MediaId + ")")

        mediaSerializer = MediaSerializer(
            media, many=True)

        mediaAdType = MediaAdType.objects.raw(
            "SELECT * FROM app_vertical_media_mediaadtype WHERE MediaId IN (" + MediaId + ")")

        mediaAdTypeSerializer = MediaAdTypeSerializer(
            mediaAdType, many=True)

        plan = MediaAdTypeImage.objects.raw(
            "SELECT * FROM app_vertical_media_mediaadtypeimage WHERE MediaAdTypeId IN (SELECT MediaAdTypeId FROM app_vertical_media_mediaadtype WHERE MediaId IN (" + MediaId + "))")
        mediaAdTypeImageSerializer = UploadMediaAdTypeImageSerializer(
            plan, many=True)

        Data = []
        Data.append(mediaSerializer.data)
        Data.append(mediaAdTypeSerializer.data)
        Data.append(mediaAdTypeImageSerializer.data)
        return JsonResponse(Data, safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


# def CURDMediaImageAPi(request, pk=""):
#     try:
#         if request.method == 'GET':
#             if not pk:
#                 mediaImaage = MediaImage.objects.all()
#                 mediaImaageSerializer = UploadMediaImageSerializer(
#                     mediaImaage, many=True)
#                 return JsonResponse(mediaImaageSerializer.data, safe=False)
#             else:
#                 mediaImaage = MediaImage.objects.get(
#                     MediaId=pk)
#                 mediaImaageSerializer = UploadMediaImageSerializer(
#                     mediaImaage)
#                 return JsonResponse(mediaImaageSerializer.data, safe=False)
#         else:
#             return JsonResponse("Invalid request", safe=False)
#     except:
#         return JsonResponse("Invalid payload", safe=False)


class MediaFileView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, pk="", ImageId="", *args, **kwargs):
        try:
            if not pk:
                file_serializer = UploadMediaSerializer(
                    data=request.data)
                if file_serializer.is_valid():
                    file_serializer.save()
                    return JsonResponse("ok", safe=False)
                else:
                    return JsonResponse("failed", safe=False)
            else:
                media = Media.objects.get(
                    MediaId=pk)
                mediaSerializer = UploadMediaSerializer(
                    media, data=request.data)
                if mediaSerializer.is_valid():
                    mediaSerializer.save()
                    return JsonResponse("ok", safe=False)
                else:
                    return JsonResponse(pk, safe=False)
        except:
            return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDMediaAdTypeForDevApi(request, pk=""):
    # try:
    mediaAdType = connection.cursor()
    mediaAdType.execute(
        "SELECT app_vertical_media_mediaadtype.MediaAdTypeId, app_vertical_media_mediaadtype.MediaAdType, app_vertical_media_mediaadtype.SiteName, app_vertical_media_mediaadtype.PackageName , app_vertical_media_mediaadtype.UnitType , app_vertical_media_mediaadtype.ShowOnOff as MediaAdUntShowOnOff, app_vertical_media_media.MediaName, app_vertical_media_media.MediaId, app_vertical_media_media.VerticalType, app_vertical_media_media.CountryEvent, app_vertical_media_media.CountryEvent , app_vertical_media_media.ShowOnOff as MediaShowOnOff , app_vertical_media_media.ViewOnlyMediaStore FROM app_vertical_media_mediaadtype JOIN app_vertical_media_media ON app_vertical_media_mediaadtype.MediaId=app_vertical_media_media.MediaId")
    field_names = [i[0] for i in mediaAdType.description]
    allMediaAdUnit = []
    for media in mediaAdType.fetchall():
        j = 0
        MediaAdUnit = {}
        for col in media:
            if j < len(field_names):
                MediaAdUnit[str(field_names[j])] = col
                j = j + 1
        allMediaAdUnit.append(MediaAdUnit)

    return JsonResponse(allMediaAdUnit, safe=False)


@csrf_exempt
def CURDMediaAdTypeApi(request, pk=""):
    # try:
    if request.method == 'GET':
        if not pk:
            mediaAdType = connection.cursor()
            mediaAdType.execute(
                "SELECT app_vertical_media_mediaadtype.MediaAdTypeId, app_vertical_media_mediaadtype.MediaAdType, app_vertical_media_mediaadtype.SiteName, app_vertical_media_mediaadtype.PackageName,app_vertical_media_media.MediaName FROM app_vertical_media_mediaadtype JOIN app_vertical_media_media ON app_vertical_media_mediaadtype.MediaId=app_vertical_media_media.MediaId AND  app_vertical_media_media.Status IN ('2','3') AND app_vertical_media_media.ShowOnOff='1'  AND app_vertical_media_mediaadtype.Status IN ('2','3') AND app_vertical_media_mediaadtype.ShowOnOff='1'")
            field_names = [i[0] for i in mediaAdType.description]
            allMediaAdUnit = []
            for media in mediaAdType.fetchall():
                j = 0
                MediaAdUnit = {}
                for col in media:
                    if j < len(field_names):
                        MediaAdUnit[str(field_names[j])] = col
                        j = j + 1
                allMediaAdUnit.append(MediaAdUnit)

            return JsonResponse(allMediaAdUnit, safe=False)
        else:

            mediaAdType = MediaAdType.objects.get(
                MediaAdTypeId=pk)
            mediaAdTypeSerializer = MediaAdTypeSerializer(
                mediaAdType)
            return JsonResponse(mediaAdTypeSerializer.data, safe=False)
    elif request.method == 'POST':
        mediaAdTypeData = JSONParser().parse(request)

        if mediaAdTypeData['Availability']:
            mediaAdTypeData['AvailabilityJSON'] = json.loads(
                mediaAdTypeData['Availability'])
        if mediaAdTypeData['NearByAdvantages']:
            mediaAdTypeData['NearByAdvantagesJSON'] = json.loads(
                mediaAdTypeData['NearByAdvantages'])
        if mediaAdTypeData['MatchingCategories']:
            mediaAdTypeData['MatchingCategoriesJSON'] = json.loads(
                mediaAdTypeData['MatchingCategories'])
        if mediaAdTypeData['AgeGroup']:
            mediaAdTypeData['AgeGroupJSON'] = json.loads(
                mediaAdTypeData['AgeGroup'])
        if mediaAdTypeData['GenderGroup']:
            mediaAdTypeData['GenderGroupJSON'] = json.loads(
                mediaAdTypeData['GenderGroup'])
        if mediaAdTypeData['IncomeGroup']:
            mediaAdTypeData['IncomeGroupJSON'] = json.loads(
                mediaAdTypeData['IncomeGroup'])
        if mediaAdTypeData['TermsAndConditions']:
            mediaAdTypeData['TermsAndConditionsNonHTML'] = cleanhtml(
                mediaAdTypeData['TermsAndConditions'])
        if mediaAdTypeData['MediaAdUnitDescription']:
            mediaAdTypeData['MediaAdUnitDescriptionNonHTML'] = cleanhtml(
                mediaAdTypeData['MediaAdUnitDescription'])
        mediaAdTypeSerializer = SaveMediaAdTypeSerializer(
            data=mediaAdTypeData)
        if mediaAdTypeSerializer.is_valid():
            mediaAdTypeSerializer.save()
            mediaAd = mediaAdTypeSerializer.data
            last_id = MediaAdType.objects.aggregate(
                Max('Identification'))['Identification__max']
            mediaAd['Identification'] = last_id+1
            mediaAdType = MediaAdType.objects.get(
                MediaAdTypeId=mediaAdTypeSerializer.data['MediaAdTypeId'])
            mediaAdTypeSerializer2 = SaveMediaAdTypeSerializer(
                mediaAdType, mediaAd)
            if mediaAdTypeSerializer2.is_valid():
                mediaAdTypeSerializer2.save()
                data=saveToRagSystemByAdUnitId(mediaAdTypeSerializer.data['MediaAdTypeId'],mediaAdTypeSerializer.data['MediaId'])
                return JsonResponse(mediaAdTypeSerializer.data['MediaAdTypeId'], safe=False)
        return JsonResponse("Failed to add", safe=False)
    elif request.method == 'PUT':
        mediaAdTypeData = JSONParser().parse(request)
        if mediaAdTypeData['Availability']:
            mediaAdTypeData['AvailabilityJSON'] = json.loads(
                mediaAdTypeData['Availability'])
        if mediaAdTypeData['NearByAdvantages']:
            mediaAdTypeData['NearByAdvantagesJSON'] = json.loads(
                mediaAdTypeData['NearByAdvantages'])
        if mediaAdTypeData['MatchingCategories']:
            mediaAdTypeData['MatchingCategoriesJSON'] = json.loads(
                mediaAdTypeData['MatchingCategories'])
        if mediaAdTypeData['AgeGroup']:
            mediaAdTypeData['AgeGroupJSON'] = json.loads(
                mediaAdTypeData['AgeGroup'])
        if mediaAdTypeData['GenderGroup']:
            mediaAdTypeData['GenderGroupJSON'] = json.loads(
                mediaAdTypeData['GenderGroup'])
        if mediaAdTypeData['IncomeGroup']:
            mediaAdTypeData['IncomeGroupJSON'] = json.loads(
                mediaAdTypeData['IncomeGroup'])
        if mediaAdTypeData['TermsAndConditions']:
            mediaAdTypeData['TermsAndConditionsNonHTML'] = cleanhtml(
                mediaAdTypeData['TermsAndConditions'])
        if mediaAdTypeData['MediaAdUnitDescription']:
            mediaAdTypeData['MediaAdUnitDescriptionNonHTML'] = cleanhtml(
                mediaAdTypeData['MediaAdUnitDescription'])
        mediaAdType = MediaAdType.objects.get(
                MediaAdTypeId=mediaAdTypeData['MediaAdTypeId'])
        mediaAdTypeSerializer = SaveMediaAdTypeSerializer(
                mediaAdType, data=mediaAdTypeData)
        if mediaAdTypeSerializer.is_valid():
            mediaAdTypeSerializer.save()
            data=saveToRagSystemByAdUnitId(mediaAdTypeSerializer.data['MediaAdTypeId'],mediaAdTypeSerializer.data['MediaId'])
            # return JsonResponse("Update successfully", safe=False)
            return JsonResponse(data, safe=False)
        
        return JsonResponse("Failed to Update", safe=False)
    elif request.method == 'DELETE':
        mediaAdType = MediaAdType.objects.get(
            MediaAdTypeId=pk)
        mediaAdType.delete()
        return JsonResponse("Deleted successfully", safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


class MediaAdTypeFileView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, pk="", *args, **kwargs):
        try:
            if not pk:
                file_serializer = UploadMediaAdTypeSerializer(
                    data=request.data)
                if file_serializer.is_valid():
                    file_serializer.save()
                    return JsonResponse("ok", safe=False)
                else:
                    return JsonResponse("failed", safe=False)
            else:
                mediaAdType = MediaAdType.objects.get(
                    MediaAdTypeId=pk)
                MediaAdType.objects.get(
                    MediaAdTypeId=pk).ThumbnailImage.delete(save=True)
                mediaAdTypeSerializer = UploadMediaAdTypeSerializer(
                    mediaAdType, data=request.data)
                if mediaAdTypeSerializer.is_valid():
                    mediaAdTypeSerializer.save()
                    return JsonResponse("ok", safe=False)
                else:
                    return JsonResponse(pk, safe=False)
        except:
            return JsonResponse("Invalid payload", safe=False)


class MediaAdTypeImageFileView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, pk="", ImageId="", *args, **kwargs):
        try:
            if not ImageId:
                file_serializer = UploadMediaAdTypeImageSerializer(
                    data=request.data)
                if file_serializer.is_valid():
                    file_serializer.save()
                    return JsonResponse("ok", safe=False)
                else:
                    return JsonResponse("failed", safe=False)
            else:
                media = MediaAdTypeImage.objects.get(
                    MediaAdTypeImageId=ImageId)
                MediaAdTypeImage.objects.get(
                    MediaAdTypeImageId=ImageId).MediaAdTypeImage.delete(save=True)
                mediaSerializer = UploadMediaAdTypeImageSerializer(
                    media, data=request.data)
                if mediaSerializer.is_valid():
                    mediaSerializer.save()
                    return JsonResponse("ok", safe=False)
                else:
                    return JsonResponse(pk, safe=False)
        except:
            return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDMediaAdTypeImageApi(request, pk=""):
    try:
        if request.method == 'GET':
            # if not pk:
            mediaAdType = MediaAdTypeImage.objects.all()
            mediaAdTypeSerializer = UploadMediaAdTypeImageSerializer(
                mediaAdType, many=True)
            return JsonResponse(mediaAdTypeSerializer.data, safe=False)
            # else:
            #     mediaAdType = MediaAdTypeImage.objects.filter(
            #         MediaAdTypeId=pk, MediaId=mediaId)
            #     mediaAdTypeSerializer = UploadMediaAdTypeImageSerializer(
            #         mediaAdType)
            #     return JsonResponse(mediaAdTypeSerializer.data, safe=False)
        elif request.method == 'DELETE':
            mediaAdTypeImage = MediaAdTypeImage.objects.get(
                MediaAdTypeImageId=pk)
            MediaAdTypeImage.objects.get(
                MediaAdTypeImageId=pk).MediaAdTypeImage.delete(save=True)
            mediaAdTypeImage.delete()
            return JsonResponse("Deleted successfully", safe=False)
        else:
            return JsonResponse("Invalid request", safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDMediaAdTypeByMediaApi(request, MediaId=""):
    # try:
    if request.method == 'GET':
        if MediaId:
            mediaAdType = MediaAdType.objects.filter(
                MediaId=MediaId)
            mediaAdTypeSerializer = MediaAdTypeSerializer(
                mediaAdType, many=True)
            return JsonResponse(mediaAdTypeSerializer.data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDMediaAdTypeByMediaIdentification(request, MediaId=""):
    # try:
    if request.method == 'GET':
        if MediaId:
            # mediaAdType = MediaAdType.objects.filter(
            #     MediaId=MediaId)
            mediaAdType = MediaAdType.objects.raw(
                "SELECT * FROM app_vertical_media_mediaadtype WHERE MediaId IN (SELECT MediaId FROM app_vertical_media_media WHERE Identification='" + MediaId + "')")
            mediaAdTypeSerializer = MediaAdTypeSerializer(
                mediaAdType, many=True)
            return JsonResponse(mediaAdTypeSerializer.data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDMediaAdTypeImageByMediaIdentification(request, MediaId=""):
    # try:
    if request.method == 'GET':
        if MediaId:
            plan = MediaAdTypeImage.objects.raw(
                "SELECT * FROM app_vertical_media_mediaadtypeimage WHERE MediaAdTypeId IN (SELECT MediaAdTypeId FROM app_vertical_media_mediaadtype WHERE MediaId IN (SELECT MediaId FROM app_vertical_media_media WHERE Identification='" + MediaId + "'))")
            planSerializer = UploadMediaAdTypeImageSerializer(
                plan, many=True)
            return JsonResponse(planSerializer.data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDMediaAdTypeImageByMediaApi(request, MediaId=""):
    # try:
    if request.method == 'GET':
        if MediaId:
            plan = MediaAdTypeImage.objects.raw(
                "SELECT * FROM app_vertical_media_mediaadtypeimage WHERE MediaAdTypeId IN (SELECT MediaAdTypeId FROM app_vertical_media_mediaadtype WHERE MediaId='" + MediaId + "')")
            planSerializer = UploadMediaAdTypeImageSerializer(
                plan, many=True)
            return JsonResponse(planSerializer.data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def DeleteAdUnitMediakit(request, pk=""):
    try:
        if request.method == 'DELETE':
            media = MediaAdType.objects.get(MediaAdTypeId=pk)
            media.AdTypeMediaKit.delete(save=True)
            mediaSerializer = SaveMediaSerializer(
                media, data=media)
            if mediaSerializer.is_valid():
                mediaSerializer.save()
            return JsonResponse("", safe=False)
        else:
            return JsonResponse("Invalid request", safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def DeleteMediakit(request, pk=""):
    try:
        if request.method == 'DELETE':
            media = Media.objects.get(MediaId=pk)
            media.MediaKit.delete(save=True)
            return JsonResponse("", safe=False)
        else:
            return JsonResponse(request.method, safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def RecommendedMediaByMedia(request):
    try:
        if request.method == 'POST':
            mediaData = JSONParser().parse(request)
            media = Media.objects.raw(
                "SELECT * FROM app_vertical_media_media WHERE MediaId IN ("+mediaData['MediaList']+")")
            mediaSerializer = MediaSerializer(
                media, many=True)
            return JsonResponse(mediaSerializer.data, safe=False)
        else:
            return JsonResponse(request.method, safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def GetRecommendedAdUnit(request):
    # try:
    if request.method == 'POST':
        mediaData = JSONParser().parse(request)
        cursor = connection.cursor()
        cursor.execute(
            "SELECT * FROM app_admin_highlyrecommendedmedia WHERE AudienceId='"+mediaData['Audience']+"' AND CountryId='"+mediaData['Country']+"'")
        MediaData = []
        for x in cursor.fetchall():
            if (x[5]):
                Media = json.loads(x[5])
            MediaAdTypeId = "''"
            for m in Media:
                MediaAdTypeId = MediaAdTypeId+",'"+m+"'"

            mediaAdType = connection.cursor()
            mediaAdType.execute(
                "SELECT app_vertical_media_mediaadtype.*,app_vertical_media_media.*,app_vertical_media_media.Identification as MediaIdentification,app_vertical_media_mediaadtype.Identification as AdUnitIdentification,app_vertical_media_media.ShowOnOff as MediaShowOnOff,app_vertical_media_mediaadtype.ShowOnOff as AdUnitShowOnOff,app_vertical_media_media.Status as MediaStatus,app_vertical_media_mediaadtype.Status as AdUnitStatus FROM app_vertical_media_mediaadtype JOIN app_vertical_media_media ON app_vertical_media_mediaadtype.MediaId=app_vertical_media_media.MediaId AND MediaAdTypeId IN (" + MediaAdTypeId + ")")
            field_names = [i[0] for i in mediaAdType.description]
            allMediaAdUnit = []
            for media in mediaAdType.fetchall():
                j = 0
                MediaAdUnit = {}
                for col in media:
                    if j < len(field_names):
                        MediaAdUnit[str(field_names[j])] = col
                        j = j + 1
                allMediaAdUnit.append(MediaAdUnit)

            category = []
            category.append(x[0])
            category.append(x[1])
            category.append(x[2])
            category.append(x[3])
            category.append(x[4])
            category.append(allMediaAdUnit)
            category.append(Media)
            MediaData.append(category)
        return JsonResponse(MediaData, safe=False)
    else:
        return JsonResponse(request.method, safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def GetRecommendedAdUnitByContryAudenc(request):
    # try:
    if request.method == 'POST':
        mediaData = JSONParser().parse(request)
        highlyrecommendedmedia = connection.cursor()
        highlyrecommendedmedia.execute(
            "SELECT * FROM app_admin_highlyrecommendedmedia WHERE AudienceId='"+mediaData['Audience']+"' AND CountryId='"+mediaData['Country']+"'")
        highlyrecommendedmedia_field_names = [
            i[0] for i in highlyrecommendedmedia.description]
        MediaData = []
        # highlyrecommendedmedia_DATA = []
        for highlyrecommendedmedia_ROW in highlyrecommendedmedia.fetchall():
            j = 0
            recommendedmedia = {}
            for col in highlyrecommendedmedia_ROW:
                if j < len(highlyrecommendedmedia_field_names):
                    recommendedmedia[str(
                        highlyrecommendedmedia_field_names[j])] = col
                    j = j + 1
            # highlyrecommendedmedia_DATA.append(recommendedmedia)
            if (highlyrecommendedmedia_ROW[5]):
                Media = json.loads(highlyrecommendedmedia_ROW[5])
            MediaAdTypeId = "''"
            for media_ROW in Media:
                MediaAdTypeId = MediaAdTypeId+",'"+media_ROW+"'"

            mediaAdType = connection.cursor()
            mediaAdType.execute(
                # "SELECT app_vertical_media_mediaadtype.*,app_vertical_media_media.*,app_vertical_media_media.Identification as MediaIdentification,app_vertical_media_mediaadtype.Identification as AdUnitIdentification,app_vertical_media_media.ShowOnOff as MediaShowOnOff,app_vertical_media_mediaadtype.ShowOnOff as AdUnitShowOnOff,app_vertical_media_media.Status as MediaStatus,app_vertical_media_mediaadtype.Status as AdUnitStatus FROM app_vertical_media_mediaadtype JOIN app_vertical_media_media ON app_vertical_media_mediaadtype.MediaId=app_vertical_media_media.MediaId  AND MediaAdTypeId IN (" + MediaAdTypeId + ")")
                "SELECT app_vertical_media_mediaadtype.*,app_vertical_media_media.*,app_vertical_media_media.Identification as MediaIdentification,app_vertical_media_mediaadtype.Identification as AdUnitIdentification,app_vertical_media_media.ShowOnOff as MediaShowOnOff,app_vertical_media_mediaadtype.ShowOnOff as AdUnitShowOnOff,app_vertical_media_media.Status as MediaStatus,app_vertical_media_mediaadtype.Status as AdUnitStatus FROM app_vertical_media_mediaadtype JOIN app_vertical_media_media ON app_vertical_media_mediaadtype.MediaId=app_vertical_media_media.MediaId and app_vertical_media_media.ShowOnOff='1' and app_vertical_media_mediaadtype.ShowOnOff='1' and app_vertical_media_mediaadtype.Status IN ('2','3') and  app_vertical_media_media.Status IN ('2','3') AND MediaAdTypeId IN (" + MediaAdTypeId + ")")
            field_names = [i[0] for i in mediaAdType.description]
            allMediaAdUnit = []
            for media in mediaAdType.fetchall():
                j = 0
                MediaAdUnit = {}
                for col in media:
                    if j < len(field_names):
                        MediaAdUnit[str(field_names[j])] = col
                        j = j + 1
                allMediaAdUnit.append(MediaAdUnit)

            category = [{}]
            # category.append(highlyrecommendedmedia_ROW[0])
            # category.append(highlyrecommendedmedia_ROW[1])
            # category.append(highlyrecommendedmedia_ROW[2])
            # category.append(highlyrecommendedmedia_ROW[3])
            # category.append(highlyrecommendedmedia_ROW[4])
            # category.append(allMediaAdUnit)
            # category.append(Media)
            recommendedmedia['AdUnit'] = allMediaAdUnit
            recommendedmedia['MediaAdTypeId'] = Media
            MediaData.append(recommendedmedia)
        return JsonResponse(MediaData, safe=False)
    else:
        return JsonResponse(request.method, safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def GetAllAdUnitForRecommendation(request):
    # try:
    if request.method == 'POST':
        mediaData = JSONParser().parse(request)
        mediaAdType = connection.cursor()
        mediaAdType.execute(
            "SELECT " +
            "app_vertical_media_mediaadtype.MediaAdTypeId, app_vertical_media_mediaadtype.MediaAdType, app_vertical_media_mediaadtype.SpecificCategory, app_vertical_media_mediaadtype.SiteName, app_vertical_media_mediaadtype.PackageName,app_vertical_media_media.MediaName,app_vertical_media_media.VerticalType " +
            "FROM app_vertical_media_mediaadtype JOIN app_vertical_media_media " +
            "ON app_vertical_media_mediaadtype.MediaId=app_vertical_media_media.MediaId  AND  " +
            "app_vertical_media_media.CountryEvent  LIKE '%%" + mediaData['CountryId'] + "%%'  AND app_vertical_media_mediaadtype.MatchingCategories  LIKE '%%" + mediaData['AudienceId'] + "%%'  ")
        field_names = [i[0] for i in mediaAdType.description]
        allMediaAdUnit = []
        for media in mediaAdType.fetchall():
            j = 0
            MediaAdUnit = {}
            for col in media:
                if j < len(field_names):
                    MediaAdUnit[str(field_names[j])] = col
                    j = j + 1
            allMediaAdUnit.append(MediaAdUnit)
        finalMedia = []
        for media in allMediaAdUnit:
            if mediaData['Type'] == 'Vertical' and media['VerticalType'] == mediaData['TypeId']:
                finalMedia.append(media)
            if mediaData['Type'] == 'SpecificCategory' and media['SpecificCategory'] == mediaData['TypeId']:
                finalMedia.append(media)

        return JsonResponse(finalMedia, safe=False)
    else:
        return JsonResponse(request.method, safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDMediaIdentification(request, VerticalId="", pk=""):
    try:
        if request.method == 'GET':
            media = Media.objects.get(
                Identification=pk)
            mediaSerializer = MediaSerializer(
                media)
            return JsonResponse(mediaSerializer.data, safe=False)
        else:
            return JsonResponse("Invalid request", safe=False)
    except:
        return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def CURDCountryByMediaAvailable(request):
    # try:
    if request.method == 'GET':
        AllMedia = connection.cursor()
        AllMedia.execute("SELECT CountryEvent FROM app_vertical_media_media WHERE MediaId IN (SELECT MediaId FROM app_vertical_media_mediaadtype WHERE (Status='2' OR Status='3') AND ShowOnOff='1' ) AND (Status='2' OR Status='3') AND ShowOnOff='1' AND ViewOnlyMediaStore = '0'")
        countryString = ""
        for media in AllMedia.fetchall():
            mediaCountry = json.loads(media[0])
            for MC in mediaCountry:
                if countryString:
                    countryString = countryString+","
                countryString = countryString + "'" + MC+"'"
        AllCountry = connection.cursor()
        AllCountry.execute(
            "SELECT * FROM app_admin_countryevent WHERE  CountryOnOFF='1' AND CountryEventId in ("+countryString+")")
        field_names = [i[0] for i in AllCountry.description]
        allCountry = []
        for media in AllCountry.fetchall():
            j = 0
            Country = {}
            for col in media:
                if j < len(field_names):
                    Country[str(field_names[j])] = col
                    j = j + 1
            allCountry.append(Country)
        # FinalCountry = []
        # for country in allCountry:
        #     count = 0
        #     for media in AllMedia.fetchall():
        #         mediaCountry = json.loads(media[0])
        #         for MC in mediaCountry:
        #             if MC == country['CountryEventId']:
        #                 count = count+1
        #     print(country['CountryEventId']+"   ---    "+str(count))
        #     if count > 0:
        #         FinalCountry.append(Country)
        return JsonResponse(allCountry, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)

# def JsonConversion(request):
#     # try:
#         if request.method == 'GET':
#             # productFaqData = JSONParser().parse(request)
#             try:
#                     filename = f"Merchant_.xlsx"
#                     # Export to Excel
#                     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#                     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#                     workbook = Workbook()
#                     worksheet = workbook.active
#                     # Write header
#                     header = list(api_data[0].keys())
#                     worksheet.append(header)
#                     # Write data
#                     for row in api_data:
#                         for key, value in row.items():
#                             if key == 'api_data' and isinstance(value, dict):
#                                 # Convert problematic fields to string if needed
#                                 for field in value:
#                                     if isinstance(value[field], (int, float)):
#                                         value[field] = str(value[field])
#                         worksheet.append(list(row.values()))
#                     workbook.save(response)
#                     return response
#             except:
#                 return JsonResponse("Invalid", safe=False)


@csrf_exempt
def GetAdunitData(request, count):
    AdunitData = []
    limit = 500
    offset = 5000
    mediaAdType = MediaAdType.objects.raw(
        "SELECT * FROM app_vertical_media_mediaadtype WHERE Status IN (2,1) and ShowOnOff=1  LIMIT "+str(limit)+" OFFSET "+str(offset)+"")

    # mediaSerializer = MediaSerializer(
    #     media, many=True)
    # mediaAdType = MediaAdType.objects.filter(
    #     Q(Q(Status=3) | Q(Status=2)), ShowOnOff=1)
    mediaAdTypeSerializer = MediaAdTypeSerializer(
        mediaAdType, many=True)
    AdunitData = []
    for adunit in mediaAdTypeSerializer.data:
        adunitData = {}
        mediaSerializer = ""
        try:
            media = Media.objects.get(Q(Q(Status=3) | Q(
                Status=2)), ShowOnOff=1, MediaId=adunit['MediaId'], CountryEvent__icontains="78fe6f9d-c707-4a4d-9609-36ceac6bfb32")
            mediaSerializer = MediaSerializer(
                media)
        except:
            p = ""
        if mediaSerializer:
            adunitData['MediaName'] = mediaSerializer.data['MediaName']
            adunitData['SiteOwner'] = mediaSerializer.data['MediaManagedBy']
            adunitData['SEOMetaTitle'] = mediaSerializer.data['SEOMetaTitle']
            adunitData['SEOMetaDescription'] = mediaSerializer.data['SEOMetaDescription']
            adunitData['SEOMetaKeyword'] = mediaSerializer.data['SEOMetaKeyword']
            adunitData['BenefitsOfAdvertising'] = mediaSerializer.data['BenefitsOfAdvertising']
            adunitData['MediaKit'] = ""
            if mediaSerializer.data['MediaKit']:
                adunitData['MediaKit'] = "https://31103116.globalmediakit.com" + \
                    str(mediaSerializer.data['MediaKit'])
            adunitData['MapLink'] = mediaSerializer.data['MapLink']
            adunitData['Latitude'] = mediaSerializer.data['Latitude']
            adunitData['Longitude'] = mediaSerializer.data['Longitude']
            adunitData['Hyperlocal'] = mediaSerializer.data['IsHyperlocal']
            adunitData['MediaAvailability'] = mediaSerializer.data['Availability']
            # adunitData['BaseURL']="https://31103116.globalmediakit.com"

            adunitData['MediaId'] = mediaSerializer.data['MediaId']
            adunitData['VerticalId'] = mediaSerializer.data['VerticalType']

            adunitData['CoverImage'] = ""
            adunitData['AdTypeMediaKit'] = ""
            if adunit['ThumbnailImage']:
                adunitData['CoverImage'] = "https://31103116.globalmediakit.com" + \
                    str(adunit['ThumbnailImage'])

            if adunit['AdTypeMediaKit']:
                adunitData['AdTypeMediaKit'] = "https://31103116.globalmediakit.com" + \
                    str(adunit['AdTypeMediaKit'])

            adunitData['AdTypeEstimateReached'] = adunit['AdTypeEstimateReached']

            adunitData['MediaAdTypeId'] = adunit['MediaAdTypeId']
            adunitData['PromoVideoLink'] = adunit['PromoVideoLink']
            adunitData['MediaAdType'] = adunit['MediaAdType']
            adunitData['AdUnitName'] = adunit['SiteName']
            adunitData['PackageName'] = adunit['PackageName']
            adunitData['Size'] = adunit['Size']
            adunitData['NoOfScreen'] = adunit['NoOfScreen']
            adunitData['AdUnitAvailability'] = adunit['Availability']
            adunitData['Currency'] = adunit['Currency']
            adunitData['MonthlyRentalRate'] = adunit['MonthlyRentalRate']
            adunitData['Discount'] = adunit['Discount']
            adunitData['NetCostBeforeTax'] = adunit['NetCostBeforeTax']
            adunitData['ProductionCost'] = adunit['ProductionCost']
            adunitData['Tax'] = adunit['Tax']
            adunitData['NetCostBeforeTax'] = adunit['NetCostBeforeTax']
            adunitData['NetCost'] = adunit['NetCost']
            adunitData['RateValidaty'] = adunit['RateValidaty']
            adunitData['RunAsOffer'] = adunit['RunAsOffer']
            adunitData['TermsAndConditions'] = adunit['TermsAndConditions']
            adunitData['LocalTaxApplied'] = adunit['IsLocalTaxApplied']
            adunitData['NeedApprovalForCreative'] = adunit['IsNeedApproval']
            adunitData['NoOfDaysRequired'] = adunit['NoOfDaysRequired']
            adunitData['RateDisplay'] = adunit['RateDisplay']
            adunitData['MediaAdUnitDescription'] = remove_html_tags(
                adunit['MediaAdUnitDescription'])

            adunitData['SpecificCategory'] = ""
            if adunit['SpecificCategory']:
                try:
                    adUnitSpecificCategory = AdUnitSpecificCategory.objects.get(
                        AdUnitSpecificCategoryId=adunit['SpecificCategory'])
                    adUnitSpecificCategorySerializer = AdUnitSpecificCategorySerializer(
                        adUnitSpecificCategory)
                    adunitData['SpecificCategory'] = adUnitSpecificCategorySerializer.data['AdUnitSpecificCategory']
                except:
                    p = ""

            adunitData['AdTypeEstimateReached'] = ""
            if adunit['AdTypeEstimateReached']:
                try:
                    estimatedReach = EstimatedReach.objects.get(
                        EstimatedReachId=adunit['AdTypeEstimateReached'])
                    estimatedReachSerializer = EstimatedReachSerializer(
                        estimatedReach)
                    adunitData['AdTypeEstimateReached'] = estimatedReachSerializer.data['EstimatedReach']
                except:
                    p = ""

            adunitData['MediaEstimateReached'] = ""
            if mediaSerializer.data['EstimateReached']:
                try:
                    estimatedReach = EstimatedReach.objects.get(
                        EstimatedReachId=mediaSerializer.data['EstimateReached'])
                    estimatedReachSerializer = EstimatedReachSerializer(
                        estimatedReach)
                    adunitData['MediaEstimateReached'] = estimatedReachSerializer.data['EstimatedReach']
                except:
                    p = ""

            try:
                adUnitCategory = AdUnitCategory.objects.get(
                    AdUnitCategoryId=adunit['AdUnitCategory'])
                adUnitCategorySerializer = AdUnitCategorySerializer(
                    adUnitCategory)
                adunitData['AdUnitCategory'] = adUnitCategorySerializer.data['AdUnitCategory']
            except:
                p = ""

            adunitData['NearByAdvantages'] = ""
            NearByAdvantagesId = json.loads(adunit['NearByAdvantages'])
            for nearByAdvantagesId in NearByAdvantagesId:
                try:
                    nearByAdvantages = NearByAdvantage.objects.get(
                        NearByAdvantageId=nearByAdvantagesId)
                    nearByAdvantagesSerializer = NearByAdvantageSerializer(
                        nearByAdvantages)
                    if adunitData['NearByAdvantages']:
                        adunitData['NearByAdvantages'] = adunitData['NearByAdvantages']+", "
                    adunitData['NearByAdvantages'] = adunitData['NearByAdvantages'] + \
                        nearByAdvantagesSerializer.data['NearByAdvantage']
                except:
                    p = ""

            adunitData['MatchingCategories'] = ""
            MatchingCategoriesId = json.loads(adunit['MatchingCategories'])
            for matchingCategoriesId in MatchingCategoriesId:
                try:
                    matchingCategories = MatchingCategory.objects.get(
                        MatchingCategoryId=matchingCategoriesId)
                    matchingCategoriesSerializer = MatchingCategorySerializer(
                        matchingCategories)
                    if adunitData['MatchingCategories']:
                        adunitData['MatchingCategories'] = adunitData['MatchingCategories']+", "
                    adunitData['MatchingCategories'] = adunitData['MatchingCategories'] + \
                        matchingCategoriesSerializer.data['MatchingCategory']
                except:
                    p = ""

            adunitData['AgeGroup'] = ""
            try:
                AgeGroupId = json.loads(adunit['AgeGroup'])
                for ageGroupId in AgeGroupId:
                    try:
                        ageGroup = AgeGroup.objects.get(
                            AgeGroupId=ageGroupId)
                        ageGroupSerializer = AgeGroupSerializer(
                            ageGroup)
                        if adunitData['AgeGroup']:
                            adunitData['AgeGroup'] = adunitData['AgeGroup']+", "
                        adunitData['AgeGroup'] = adunitData['AgeGroup'] + \
                            ageGroupSerializer.data['AgeGroup']
                    except:
                        p = ""
            except:
                p = ""

            adunitData['GenderGroup'] = ""
            try:
                GenderGroupId = json.loads(adunit['GenderGroup'])
                for genderGroupId in GenderGroupId:
                    try:
                        genderGroup = GenderGroup.objects.get(
                            GenderGroupId=genderGroupId)
                        genderGroupSerializer = GenderGroupSerializer(
                            genderGroup)
                        if adunitData['GenderGroup']:
                            adunitData['GenderGroup'] = adunitData['GenderGroup']+", "
                        adunitData['GenderGroup'] = adunitData['GenderGroup'] + \
                            genderGroupSerializer.data['GenderGroup']
                    except:
                        p = ""
            except:
                p = ""

            adunitData['IncomeGroup'] = ""
            try:
                IncomeGroupId = json.loads(adunit['IncomeGroup'])
                for incomeGroupId in IncomeGroupId:
                    try:
                        incomeGroup = IncomeGroup.objects.get(
                            IncomeGroupId=incomeGroupId)
                        incomeGroupSerializer = IncomeGroupSerializer(
                            incomeGroup)
                        if adunitData['IncomeGroup']:
                            adunitData['IncomeGroup'] = adunitData['IncomeGroup']+", "
                        adunitData['IncomeGroup'] = adunitData['IncomeGroup'] + \
                            incomeGroupSerializer.data['IncomeGroup']
                    except:
                        p = ""
            except:
                p = ""

            adunitData['NationalityCommunity'] = ""
            if adunit['NationalityCommunity']:
                try:
                    NationalityCommunityId = json.loads(
                        adunit['NationalityCommunity'])
                    for nationalityCommunityId in NationalityCommunityId:
                        nationalityCommunity = NationalityCommunity.objects.get(
                            NationalityCommunityId=nationalityCommunityId)
                        nationalityCommunitySerializer = NationalityCommunitySerializer(
                            nationalityCommunity)
                        if adunitData['NationalityCommunity']:
                            adunitData['NationalityCommunity'] = adunitData['NationalityCommunity']+", "
                        adunitData['NationalityCommunity'] = adunitData['NationalityCommunity'] + \
                            nationalityCommunitySerializer.data['NationalityCommunity']
                except:
                    try:
                        NationalityCommunityId = adunit['NationalityCommunity']
                        nationalityCommunity = NationalityCommunity.objects.get(
                            NationalityCommunityId=NationalityCommunityId)
                        nationalityCommunitySerializer = NationalityCommunitySerializer(
                            nationalityCommunity)
                        adunitData['NationalityCommunity'] = nationalityCommunitySerializer.data['NationalityCommunity']
                    except:
                        p = ""

            adunitData['Country'] = ''
            try:
                CountryEventId = json.loads(
                    mediaSerializer.data['CountryEvent'])
                for country in CountryEventId:
                    try:
                        country = CountryEvent.objects.get(
                            CountryEventId=country)
                        countrySerializer = CountryEventSerializer(
                            country)
                        if adunitData['Country']:
                            adunitData['Country'] = adunitData['Country']+", "
                        adunitData['Country'] = adunitData['Country'] + \
                            countrySerializer.data['CountryEventName']
                    except:
                        p = ""
            except:
                p = ""

            adunitData['CityRegion'] = ''
            try:
                CityRegionId = json.loads(mediaSerializer.data['CityRegion'])
                for cityRegionid in CityRegionId:
                    try:
                        cityRegion = CityRegion.objects.get(
                            CityId=cityRegionid)
                        cityRegionSerializer = CityRegionSerializer(
                            cityRegion)
                        if adunitData['CityRegion']:
                            adunitData['CityRegion'] = adunitData['CityRegion']+", "
                        adunitData['CityRegion'] = adunitData['CityRegion'] + \
                            cityRegionSerializer.data['CityRegionName']
                    except:
                        p = ""
            except:
                p = ""

            adunitData['Language'] = ''
            try:
                LanguageId = json.loads(mediaSerializer.data['Language'])
                for languageId in LanguageId:
                    try:
                        language = Language.objects.get(
                            LanguageId=languageId)
                        languageSerializer = LanguageSerializer(
                            language)
                        if adunitData['Language']:
                            adunitData['Language'] = adunitData['Language']+", "
                        adunitData['Language'] = adunitData['Language'] + \
                            languageSerializer.data['Language']
                    except:
                        p = ""
            except:
                p = ""

            adunitData['VerticalName'] = ""
            if mediaSerializer.data['VerticalType']:
                try:
                    verticalCategory = Vertical.objects.get(
                        VerticalId=mediaSerializer.data['VerticalType'])
                    verticalCategorySerializer = VerticalSerializer(
                        verticalCategory)
                    adunitData['VerticalName'] = verticalCategorySerializer.data['VerticalName']
                except:
                    p = ""

            adunitData['MediaCategory'] = ""
            if mediaSerializer.data['MediaCategory']:
                try:
                    mediaCategory = MediaCategory.objects.get(
                        MediaCategoryId=mediaSerializer.data['MediaCategory'])
                    mediaCategorySerializer = MediaCategorySerializer(
                        mediaCategory)
                    adunitData['MediaCategory'] = mediaCategorySerializer.data['MediaCategory']
                except:
                    p = ""

            adunitData['LocationType'] = ""
            try:
                locationType = LocationType.objects.get(
                    LocationTypeId=adunit['LocationType'])
                locationTypeSerializer = LocationTypeSerializer(locationType)
                adunitData['LocationType'] = locationTypeSerializer.data['LocationType']
            except:
                p = ""

            adunitData['PublisherName'] = ""
            if mediaSerializer.data['PublisherId']:
                try:
                    publisherRegister = PublisherRegister.objects.get(
                        PublisherId=mediaSerializer.data['PublisherId'])
                    publisherRegisterSerializer = PublisherRegisterSerializer(
                        publisherRegister)
                    adunitData['PublisherName'] = publisherRegisterSerializer.data['MediaName'] + \
                        " - " + publisherRegisterSerializer.data['CompanyName']
                except:
                    p = ""

            adunitData['MediaURL'] = "https://globalmediakit.com/media/media-detail/"+adunitData['MediaName']+"/Location/" + \
                adunitData['VerticalName']+"/"+adunitData['MediaId']+"/" + \
                adunitData['VerticalId']+"/CountryEventName/" + \
                adunitData['MediaAdTypeId']+""

            AdunitData.append(adunitData)
        # except:
        #     p = ""

    # filename = "adunit-"+str(count)+".xlsx"
    # # Export to Excel
    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = f'attachment; filename="{filename}"'
    # workbook = Workbook()
    # worksheet = workbook.active
    # # Write header
    # header = list(AdunitData[0].keys())
    # worksheet.append(header)
    # # Write data
    # for row in AdunitData:
    #     for key, value in row.items():
    #         if key == 'api_data' and isinstance(value, dict):
    #             # Convert problematic fields to string if needed
    #             for field in value:
    #                 if isinstance(value[field], (int, float)):
    #                     value[field] = str(value[field])
    #     worksheet.append(list(row.values()))
    # workbook.save('media/estimate/excel/'+filename)
    return JsonResponse(AdunitData, safe=False)


def remove_html_tags(html):
    soup = BeautifulSoup(html, 'html.parser')
    return soup.get_text()


@csrf_exempt
def AIRecommendedAdUnitApi(request):
    mediaData = JSONParser().parse(request)
    CountryId = mediaData['Country']
    Limit = mediaData['Limit']
    offset = mediaData['Offset']
    MediaCount = mediaData['MediaCount']
    City = mediaData['City']
    AIRecommendedAdUnit = mediaData['AIRecommendedAdUnit']
    Language = mediaData['Language']
    CampaignType = mediaData['CampaignType']
    VerticalId = mediaData['VerticalNSpecificCategoryId']
    if request.method == 'POST':
        Data = []
        AdUnitId = ""
        if len(AIRecommendedAdUnit):
            sub_query = "SELECT MediaId FROM gmk_db.app_vertical_media_mediaadtype where "
            for id in AIRecommendedAdUnit:

                AdUnitId = AdUnitId+",'"+id+"'"
                if sub_query:
                    sub_query = sub_query + " OR "
                sub_query = sub_query+" MediaAdTypeId='" + id+"' "
            query = query+"and  MediaId IN (" + sub_query + " )"
        media = Media.objects.raw(
            "SELECT * FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' " + query + " ")

        mediaSerializer = MediaSerializer(
            media, many=True)

        MediaId = "''"
        for x in mediaSerializer.data:
            MediaId = MediaId+",'"+x['MediaId']+"'"

        mediaAdType = MediaAdType.objects.raw(
            "SELECT * FROM app_vertical_media_mediaadtype WHERE MediaAdTypeId IN (" + AdUnitId + ")")

        mediaAdTypeSerializer = MediaAdTypeSerializer(
            mediaAdType, many=True)

        Data.append(mediaSerializer.data)
        Data.append(mediaAdTypeSerializer.data)

        return JsonResponse(Data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)


@csrf_exempt
def aiMediaForViewAllApi(request):
    mediaData = JSONParser().parse(request)
    AIRecommendedAdUnit = mediaData['AIRecommendedAdUnit']
    CountryId = mediaData['Country']
    Limit = mediaData['Limit']
    offset = mediaData['Offset']
    MediaCount = mediaData['MediaCount']
    City = mediaData['City']
    Language = mediaData['Language']
    CampaignType = mediaData['CampaignType']
    VerticalId = mediaData['VerticalNSpecificCategoryId']
    if request.method == 'POST':
        Data = []
        VerticalId = mediaData['VerticalNSpecificCategoryId']
        query = ""
        if len(CountryId):
            sub_query = ""
            for id in CountryId:
                if sub_query:
                    sub_query = sub_query + " OR "
                sub_query = sub_query + \
                    " JSON_CONTAINS(CountryEventJSON, '"+'"' + id+'"'+"') "
            query = query+" AND " + sub_query
        if len(VerticalId):
            sub_query = ""
            for id in VerticalId:
                if sub_query:
                    sub_query = sub_query + " OR "
                sub_query = sub_query+" VerticalType = '" + id + "' "
            query = query+" AND (" + sub_query + ")"

        AdUnitId = ""
        for media in AIRecommendedAdUnit:
            if AdUnitId:
                AdUnitId = AdUnitId+","
            AdUnitId = AdUnitId+"'"+media+"'"
        if MediaCount:
            Count = connection.cursor()
            Count.execute(
                "SELECT COUNT(MediaId) as Count FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' " + query)
            MediaCount = Count.fetchall()[0][0]

        media = Media.objects.raw(
            "SELECT * FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' " + query + "  LIMIT "+offset+", "+Limit+"")

        mediaSerializer = MediaSerializer(
            media, many=True)
        MediaId = "''"
        for x in mediaSerializer.data:
            MediaId = MediaId+",'"+x['MediaId']+"'"

        mediaAdType = MediaAdType.objects.raw(
            "SELECT * FROM app_vertical_media_mediaadtype WHERE (MediaId IN (" + MediaId + ") and  MediaId NOT IN (SELECT MediaId FROM app_vertical_media_mediaadtype WHERE MediaAdTypeId IN (" + AdUnitId + ")) )")

        mediaAdTypeSerializer = MediaAdTypeSerializer(
            mediaAdType, many=True)


        mediaDetail=mediaSerializer.data
        mediaAdTypeDetail=mediaAdTypeSerializer.data

        if offset==0 or offset=="0":
            if AdUnitId:
                
                media = Media.objects.raw(
                    "SELECT * FROM app_vertical_media_media WHERE ViewOnlyMediaStore='0' AND Status IN ('2','3') AND ShowOnOff='1' and MediaId IN (SELECT MediaId FROM app_vertical_media_mediaadtype WHERE MediaAdTypeId IN (" + AdUnitId + ")) ")

                mediaSerializer = MediaSerializer(
                    media, many=True)

                MediaId = "''"
                for x in mediaSerializer.data:
                    MediaId = MediaId+",'"+x['MediaId']+"'"

                mediaAdType = MediaAdType.objects.raw(
                    "SELECT * FROM app_vertical_media_mediaadtype WHERE MediaAdTypeId IN (" + AdUnitId + ")")

                mediaAdTypeSerializer = MediaAdTypeSerializer(
                    mediaAdType, many=True)
                for m in mediaSerializer.data:
                    mediaDetail.append(m)
                for m in mediaAdTypeSerializer.data:
                    mediaAdTypeDetail.append(m)
        Data.append(mediaDetail)
        Data.append(mediaAdTypeDetail)
        Data.append(MediaCount)

        return JsonResponse(Data, safe=False)
    else:
        return JsonResponse("Invalid request", safe=False)
    # except:
    #     return JsonResponse("Invalid payload", safe=False)
# from django.db import connection
# import pandas as pd
# from io import BytesIO
# @csrf_exempt
# def MediaAdTypeWithMediaViewExcel(request, pk=""):
#     # try:
#         if request.method == 'GET':
#             with connection.cursor() as cursor:
#                 cursor.execute("SELECT MediaId,MediaAdTypeId,MediaName,PackageName,SiteName,PriceRangeNote,CountryEventJSON,CityRegionJSON,VerticalTypeName,RunAsOffer,NetCost FROM MediaAdTypeWithMediaView;")
#                 media_mediaadtypeData = []
#                 countries = CountryEventSerializer(CountryEvent.objects.all(),many=True).data
#                 city_regions = CityRegionSerializer(CityRegion.objects.all(),many=True).data
#                 for media in cursor.fetchall():
#                     # Convert the tuple to a list to allow modifications
#                     media_list = list(media)
                    
#                     # Process CountryEventJSON (index 6)
#                     country_data = ""
#                     for id in (media_list[6] or []):
#                         for data in countries:
#                             if data['CountryEventId'] == id:
#                                 country_data = country_data + ", " + data['CountryEventName']
#                     media_list[6] = country_data  # Update the list

#                     # Process CityRegionJSON (index 7)
#                     city_data = ""
#                     for id in (media_list[7] or []):
#                         for data in city_regions:
#                             if data['CityId'] == id:
#                                 city_data = city_data + ", " + data['CityRegionName']
#                     media_list[7] = city_data  # Update the list

#                     # Append the modified list to the result list
#                     media_mediaadtypeData.append(media_list)

#                 # return JsonResponse(media_mediaadtypeData, safe=False)
#                 columns = [col[0] for col in cursor.description]
#                 data = [dict(zip(columns, row)) for row in media_mediaadtypeData]
#                 df = pd.DataFrame(data)
#                 output = BytesIO()
#                 with pd.ExcelWriter(output, engine='openpyxl') as writer:
#                     df.to_excel(writer, index=False)
               
               
#                 output.seek(0)
               
#                 response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#                 response['Content-Disposition'] = f'attachment; filename="data.xlsx"'
#                 return response
 
#     # except:
#     #     return JsonResponse("", safe=False)
from django.http import StreamingHttpResponse
from openpyxl import Workbook
from io import BytesIO
import json
import logging

# Set up a logger
logger = logging.getLogger(__name__)

def download_excel_view(request):
    try:
        media_mediaadtype = connection.cursor()
        media_mediaadtype.execute(
            "SELECT * FROM app_vertical_media_media JOIN app_vertical_media_mediaadtype ON `app_vertical_media_media`.`MediaId` = `app_vertical_media_mediaadtype`.`MediaId` AND `app_vertical_media_mediaadtype`.`Status` in ('2','3') AND `app_vertical_media_mediaadtype`.`ShowOnOff` = '1' AND `app_vertical_media_media`.`Status` in ('2','3') AND `app_vertical_media_media`.`ShowOnOff` = '1'  AND (`app_vertical_media_mediaadtype`.`IncomeGroup` LIKE '%%5c4963da-4d4c-49d7-a1ff-d5ae91c50fa1%' OR `app_vertical_media_mediaadtype`.`IncomeGroup` LIKE '%%bef5c4e8-106a-4088-95a1-9d666670afc6%%');"
        )

        field_names = [i[0] for i in media_mediaadtype.description]
        field_names.append('Units')  # Add Units to the field names
        media_mediaadtypeData = media_mediaadtype.fetchall()

        countries = {c['CountryEventId']: c['CountryEventName'] for c in CountryEventSerializer(CountryEvent.objects.all(), many=True).data}
        incomeGroup = {c['IncomeGroupId']: c['IncomeGroup'] for c in IncomeGroupSerializer(IncomeGroup.objects.all(), many=True).data}
        languages = {c['LanguageId']: c['Language'] for c in LanguageSerializer(Language.objects.all(), many=True).data}
        city_regions = {c['CityId']: c['CityRegionName'] for c in CityRegionSerializer(CityRegion.objects.all(), many=True).data}
        verticals = {c['VerticalId']: c['VerticalName'] for c in VerticalSerializer(Vertical.objects.filter(RefId=""), many=True).data}
        estimated_reaches = {c['EstimatedReachId']: c['EstimatedReach'] for c in EstimatedReachSerializer(EstimatedReach.objects.all(), many=True).data}
        adunit_sizes = {c['AdunitSizeId']: c['AdunitSize'] for c in AdunitSizeSerializer(AdunitSize.objects.all(), many=True).data}
        adunit_positions = {c['AdunitPositionId']: c['AdunitPosition'] for c in AdunitPositionSerializer(AdunitPosition.objects.all(), many=True).data}
        adunit_categories = {c['AdUnitCategoryId']: c['AdUnitCategory'] for c in AdUnitCategorySerializer(AdUnitCategory.objects.all(), many=True).data}
        ad_formats = {c['AdFormatId']: c['AdFormat'] for c in AdFormatSerializer(AdFormat.objects.all(), many=True).data}
        adunit_specific_categories = {c['AdUnitSpecificCategoryId']: c['AdUnitSpecificCategory'] for c in AdUnitSpecificCategorySerializer(AdUnitSpecificCategory.objects.all(), many=True).data}

        def get_units(vertical_name, media_ad_type):
            if media_ad_type == 'Package':
                return 'Package'
            if vertical_name == 'Out of Home':
                return 'Monthly'
            if vertical_name == 'Radio':
                return 'Spot'
            if vertical_name == 'Digital':
                return 'CPM'
            if vertical_name == 'Newspaper':
                return 'Ad Rate'
            if vertical_name == 'Magazine':
                return 'Ad Rate'
            if vertical_name == 'Television':
                return 'Spot'
            if vertical_name == 'Media Service':
                return 'Service'
            return ''  # Default empty if no condition is met

        # Create Excel file using openpyxl
        def generate_excel():
            wb = Workbook()
            ws = wb.active
            ws.title = "Media Data"
            
            # Write header
            ws.append(field_names)

            # Write data rows
            for media in media_mediaadtypeData:
                media_dict = dict(zip(field_names[:-1], media))  # Exclude 'Units' from the initial dict creation
                city_region_data = [city_regions.get(id, "") for id in json.loads(media_dict.get('CityRegion', '[]'))]
                language_data = [languages.get(id, "") for id in json.loads(media_dict.get('Language', '[]'))]
                country_data = [countries.get(id, "") for id in json.loads(media_dict.get('CountryEvent', '[]'))]
                incomeGroup_data = [incomeGroup.get(id, "") for id in json.loads(media_dict.get('IncomeGroup', '[]'))]
                
                media_dict.update({
                    'VerticalType': verticals.get(media_dict.get('VerticalType', ''), ""),
                    'EstimateReached': estimated_reaches.get(media_dict.get('EstimateReached', ''), ""),
                    'AdTypeEstimateReached': estimated_reaches.get(media_dict.get('AdTypeEstimateReached', ''), ""),
                    'Size': adunit_sizes.get(media_dict.get('Size', ''), ""),
                    'Position': adunit_positions.get(media_dict.get('Position', ''), ""),
                    'AdUnitCategory': adunit_categories.get(media_dict.get('AdUnitCategory', ''), ""),
                    'AdFormat': ad_formats.get(media_dict.get('AdFormat', ''), ""),
                    'SpecificCategory': adunit_specific_categories.get(media_dict.get('SpecificCategory', ''), ""),
                    'PrimaryLanguage': languages.get(media_dict.get('PrimaryLanguage', ''), ""),
                })

                # Calculate Units based on conditions
                units = get_units(media_dict.get('VerticalName', ''), media_dict.get('MediaAdType', ''))

                # Update dictionary with JSON data and Units
                media_dict.update({
                    'CountryEventJSON': ', '.join(country_data),  # Convert list to comma-separated string
                    'CityRegionJSON': ', '.join(city_region_data),  # Convert list to comma-separated string
                    'Language': ', '.join(language_data),  # Convert list to comma-separated string
                    'IncomeGroup': ', '.join(incomeGroup_data),  # Convert list to comma-separated string
                    'Units': units,  # Add Units value to the dict
                })

                # Write row data
                ws.append([media_dict.get(col, "") for col in field_names])

            # Save Excel to a bytes buffer
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        # Generate response with Excel file
        response = StreamingHttpResponse(generate_excel(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="media_data.xlsx"'
        return response

    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return StreamingHttpResponse(f"An error occurred: {str(e)}", status=500)



from django.http import JsonResponse
import pandas as pd
import json

@csrf_exempt

def getMediaData(request):
    # try:
        media_mediaadtype = connection.cursor()
        media_mediaadtype.execute(
            # "SELECT * FROM app_vertical_media_media JOIN app_vertical_media_mediaadtype ON `app_vertical_media_media`.`MediaId` = `app_vertical_media_mediaadtype`.`MediaId` AND `app_vertical_media_mediaadtype`.`Status` in ('2','3') AND `app_vertical_media_mediaadtype`.`ShowOnOff` = '1' AND `app_vertical_media_media`.`Status` in ('2','3') AND `app_vertical_media_media`.`ShowOnOff` = '1'  AND (`app_vertical_media_mediaadtype`.`IncomeGroup` LIKE '%%5c4963da-4d4c-49d7-a1ff-d5ae91c50fa1%' OR `app_vertical_media_mediaadtype`.`IncomeGroup` LIKE '%%bef5c4e8-106a-4088-95a1-9d666670afc6%%');"
            """SELECT * FROM app_vertical_media_media JOIN app_vertical_media_mediaadtype ON 
            `app_vertical_media_media`.`MediaId` = `app_vertical_media_mediaadtype`.`MediaId` AND 
            `app_vertical_media_mediaadtype`.`Status` in ('2','3') AND `app_vertical_media_mediaadtype`.`ShowOnOff` = '1' AND 
            `app_vertical_media_media`.`Status` in ('2','3') AND `app_vertical_media_media`.`ShowOnOff` = '1' 
            and app_vertical_media_media.VerticalType IN ('1166b5ee-a0a1-40b6-88b6-2713e0a058de', 
                              '4c21a24d-01eb-4995-b96f-973a58e72d77', 
                              '6f71ac89-03c9-4c99-8e54-d43fefaab1ef', 
                              'b195a568-0cc6-4a16-a552-761e795e854f')
    AND app_vertical_media_media.MediaCategory IN ('5ca8013a-97f9-431d-b62c-8640348b4ada', 
                             '7059141e-ae3d-420b-8952-0712ce30d912', 
                             '89777614-7cf3-4168-9705-950662568bac', 
                             'c37015bb-3a67-4808-ab7e-79186020f565')
            ;"""
        )

        field_names = [i[0] for i in media_mediaadtype.description]
        media_mediaadtypeData = media_mediaadtype.fetchall()

        matching_category = {c['MatchingCategoryId']: c['MatchingCategory'] for c in MatchingCategorySerializer(MatchingCategory.objects.all(), many=True).data}
        countries = {c['CountryEventId']: c['CountryEventName'] for c in CountryEventSerializer(CountryEvent.objects.all(), many=True).data}
        incomeGroup = {c['IncomeGroupId']: c['IncomeGroup'] for c in IncomeGroupSerializer(IncomeGroup.objects.all(), many=True).data}
        languages = {c['LanguageId']: c['Language'] for c in LanguageSerializer(Language.objects.all(), many=True).data}
        city_regions = {c['CityId']: c['CityRegionName'] for c in CityRegionSerializer(CityRegion.objects.all(), many=True).data}
        verticals = {c['VerticalId']: c['VerticalName'] for c in VerticalSerializer(Vertical.objects.filter(RefId=""), many=True).data}
        estimated_reaches = {c['EstimatedReachId']: c['EstimatedReach'] for c in EstimatedReachSerializer(EstimatedReach.objects.all(), many=True).data}
        adunit_sizes = {c['AdunitSizeId']: c['AdunitSize'] for c in AdunitSizeSerializer(AdunitSize.objects.all(), many=True).data}
        adunit_positions = {c['AdunitPositionId']: c['AdunitPosition'] for c in AdunitPositionSerializer(AdunitPosition.objects.all(), many=True).data}
        adunit_categories = {c['AdUnitCategoryId']: c['AdUnitCategory'] for c in AdUnitCategorySerializer(AdUnitCategory.objects.all(), many=True).data}
        ad_formats = {c['AdFormatId']: c['AdFormat'] for c in AdFormatSerializer(AdFormat.objects.all(), many=True).data}
        adunit_specific_categories = {c['AdUnitSpecificCategoryId']: c['AdUnitSpecificCategory'] for c in AdUnitSpecificCategorySerializer(AdUnitSpecificCategory.objects.all(), many=True).data}
        adunit_category = {c['AdUnitCategoryId']: c['AdUnitCategory'] for c in AdUnitCategorySerializer(AdUnitCategory.objects.all(), many=True).data}

        def get_units(vertical_name, media_ad_type):
            if media_ad_type == 'Package':
                return 'Package'
            if vertical_name == 'Out of Home':
                return 'Monthly'
            if vertical_name == 'Radio':
                return 'Spot'
            if vertical_name == 'Digital':
                return 'CPM'
            if vertical_name == 'Newspaper':
                return 'Ad Rate'
            if vertical_name == 'Magazine':
                return 'Ad Rate'
            if vertical_name == 'Television':
                return 'Spot'
            if vertical_name == 'Media Service':
                return 'Service'
            return ''  # Default empty if no condition is met

        # Prepare JSON response data
        data_list = []
        country_media_data = {}
        for media in media_mediaadtypeData:
            media_dict = dict(zip(field_names, media))
            city_region_data = [city_regions.get(id, "") for id in json.loads(media_dict.get('CityRegion', '[]'))]
            language_data = [languages.get(id, "") for id in json.loads(media_dict.get('Language', '[]'))]
            country_data = [countries.get(id, "") for id in json.loads(media_dict.get('CountryEvent', '[]'))]
            incomeGroup_data = [incomeGroup.get(id, "") for id in json.loads(media_dict.get('IncomeGroup', '[]'))]
            MatchingCategories_data = [matching_category.get(id, "") for id in json.loads(media_dict.get('MatchingCategories', '[]'))]

            media_dict.update({
                'VerticalType': verticals.get(media_dict.get('VerticalType', ''), ""),
                'EstimateReached': estimated_reaches.get(media_dict.get('EstimateReached', ''), ""),
                'AdTypeEstimateReached': estimated_reaches.get(media_dict.get('AdTypeEstimateReached', ''), ""),
                'Size': adunit_sizes.get(media_dict.get('Size', ''), ""),
                'Position': adunit_positions.get(media_dict.get('Position', ''), ""),
                'AdUnitCategory': adunit_categories.get(media_dict.get('AdUnitCategory', ''), ""),
                'AdFormat': ad_formats.get(media_dict.get('AdFormat', ''), ""),
                'SpecificCategory': adunit_specific_categories.get(media_dict.get('SpecificCategory', ''), ""),
                'PrimaryLanguage': languages.get(media_dict.get('PrimaryLanguage', ''), ""),
                'MediaCategory': adunit_category.get(media_dict.get('MediaCategory', ''), ""),
                'CountryEventJSON': ', '.join(country_data),
                'CityRegionJSON': ', '.join(city_region_data),
                'Language': ', '.join(language_data),
                'IncomeGroup': ', '.join(incomeGroup_data),
                'MatchingCategories': ', '.join(MatchingCategories_data),
                'Units': get_units(media_dict.get('VerticalName', ''), media_dict.get('MediaAdType', '')),
            })
            for country in country_data:
                if country not in country_media_data:
                    country_media_data[country] = []  # Create a list for this country if it doesn't exist
                country_media_data[country].append(media_dict)
            # return JsonResponse(country_media_data, safe=False)
            # data_list.append(media_dict)

        # Return JSON response

        # df = pd.DataFrame(data_list)

        # Create an Excel file
        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine="xlsxwriter") as writer:
            for country, data_list in country_media_data.items():
                # Convert the data list to a DataFrame
                df = pd.DataFrame(data_list)
                
                # Use country as the sheet name (truncate if > 31 chars)
                sheet_name = country[:31]
                
                # Write the DataFrame to the sheet
                df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Finalize and seek to the beginning of the BytesIO buffer
        excel_file.seek(0)

        # Serve as a response
        response = HttpResponse(
            excel_file,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=media_data_by_country.xlsx"

        return response
    # except Exception as e:
    #     logger.error(f"An error occurred: {str(e)}")
    #     return JsonResponse({'error': f"An error occurred: {str(e)}"}, status=500)
